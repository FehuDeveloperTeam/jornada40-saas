from rest_framework.decorators import api_view, permission_classes, action, throttle_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from rest_framework import viewsets
from rest_framework.exceptions import ValidationError
from rest_framework.throttling import AnonRateThrottle


class LoginRateThrottle(AnonRateThrottle):
    scope = 'login'

class RegisterRateThrottle(AnonRateThrottle):
    scope = 'register'

class PasswordResetRateThrottle(AnonRateThrottle):
    scope = 'password_reset'
from django.contrib.auth.models import User
from django.db import transaction, IntegrityError
from django.http import HttpResponse
from django.template.loader import render_to_string, get_template
from .models import Plan, Suscripcion, Cliente
from .serializers import PlanSerializer
from django.contrib.auth.forms import PasswordResetForm
from xhtml2pdf import pisa
from django.conf import settings
from django.utils import timezone
import datetime
import io
import zipfile
import re
import math
import random
import string
from num2words import num2words
import hmac
import logging
from html import escape as _esc

logger = logging.getLogger(__name__)
import pandas as pd
import traceback
import urllib.parse
from django.db.models import Max, Sum, Exists, OuterRef
from django.core.files.base import ContentFile
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .models import Plan, Cliente, Empresa, Empleado, Contrato, AnexoContrato, DocumentoLegal, Liquidacion, SolicitudFirma, OTPFirma, VacacionEmpleado, Finiquito
from .serializers import EmpresaSerializer, EmpleadoSerializer, ContratoSerializer, AnexoContratoSerializer, DocumentoLegalSerializer, LiquidacionSerializer, SolicitudFirmaSerializer, FiniquitoSerializer
from . import b2_client
from django.core.mail import EmailMultiAlternatives
import uuid as uuid_mod

# ==========================================
# UTILIDADES DE RUT (VALIDACIÓN Y FORMATO)
# ==========================================
def limpiar_rut(rut):
    return re.sub(r'[^0-9kK]', '', str(rut)).upper()

def formatear_rut(rut):
    rut_limpio = limpiar_rut(rut)
    if len(rut_limpio) < 2:
        return rut
    cuerpo = rut_limpio[:-1]
    dv = rut_limpio[-1]
    try:
        cuerpo_con_puntos = "{:,}".format(int(cuerpo)).replace(',', '.')
    except ValueError:
        return rut
    return f"{cuerpo_con_puntos}-{dv}"

def validar_rut(rut):
    rut_limpio = limpiar_rut(rut)
    if len(rut_limpio) < 2:
        return False
    cuerpo = rut_limpio[:-1]
    dv_ingresado = rut_limpio[-1]

    try:
        cuerpo_int = int(cuerpo)
    except ValueError:
        return False

    suma = 0
    multiplo = 2
    for d in reversed(cuerpo):
        suma += int(d) * multiplo
        multiplo += 1
        if multiplo == 8:
            multiplo = 2
    
    resto = suma % 11
    dv_esperado = 11 - resto
    
    if dv_esperado == 11:
        dv_calculado = '0'
    elif dv_esperado == 10:
        dv_calculado = 'K'
    else:
        dv_calculado = str(dv_esperado)
        
    return dv_ingresado == dv_calculado

# ==========================================
# TRADUCTOR INTELIGENTE DE FECHAS EXCEL
# ==========================================
def estandarizar_fecha(fecha_valor):
    # 1. Manejo de nulos (incluyendo nulos de Pandas)
    if not fecha_valor or pd.isna(fecha_valor):
        return None
    
    # 2. Si Pandas ya lo parseó correctamente como objeto datetime/date
    if isinstance(fecha_valor, (datetime.datetime, datetime.date)):
        return fecha_valor.date() if isinstance(fecha_valor, datetime.datetime) else fecha_valor

    fecha_str = str(fecha_valor).strip()
    
    # 3. Si viene como número de serie de Excel
    try:
        serial = float(fecha_str)
        base = datetime.datetime(1899, 12, 30)
        return (base + datetime.timedelta(days=serial)).date()
    except ValueError:
        pass

    # 4. Formatos estrictos chilenos (Día, Mes, Año) + ISO estándar de BD
    formatos_chilenos = [
        '%d-%m-%Y', '%d/%m/%Y', '%d.%m.%Y',
        '%d-%m-%y', '%d/%m/%y', '%d.%m.%y',
        '%Y-%m-%d'
    ]
    
    for fmt in formatos_chilenos:
        try:
            dt = datetime.datetime.strptime(fecha_str, fmt).date()
            # Ajuste para años de 2 dígitos (ej: 92 -> 1992 en vez de 2092)
            if dt.year > datetime.date.today().year + 10:
                dt = dt.replace(year=dt.year - 100)
            return dt
        except ValueError:
            continue

    return None


def _plan_activo(user):
    """Devuelve el objeto Plan activo del usuario, o None si no tiene plan."""
    cliente = getattr(user, 'perfil_cliente', None)
    if not cliente:
        return None
    plan = cliente.plan
    if not plan:
        try:
            suscripcion = cliente.suscripcion_activa
            if suscripcion.estado in ('ACTIVE', 'TRIAL', 'PAST_DUE'):
                plan = suscripcion.plan
        except Exception:
            pass
    return plan


def _nivel_plan(user) -> int:
    """Retorna el nivel del plan activo: 1=Semilla, 2=Starter, 3=Pyme, 4=Corporativo.
    Sin plan asignado se asume nivel 1 (Semilla)."""
    plan = _plan_activo(user)
    return plan.nivel if plan else 1


def _plan_permite(user, nivel_min: int) -> bool:
    """True si el plan del usuario tiene nivel >= nivel_min."""
    return _nivel_plan(user) >= nivel_min


def _es_plan_semilla(user) -> bool:
    """True si el usuario no tiene plan activo o su plan es nivel 1 (Semilla)."""
    return _nivel_plan(user) == 1


def _html_a_pdf_bytes(html_string: str, nombre_doc: str) -> bytes:
    """Convierte HTML a bytes PDF con xhtml2pdf. Lanza excepción si falla."""
    resultado = io.BytesIO()
    status = pisa.pisaDocument(io.BytesIO(html_string.encode("UTF-8")), resultado)
    if status.err:
        raise Exception(f"Error generando PDF '{nombre_doc}'.")
    pdf_bytes = resultado.getvalue()
    if not pdf_bytes:
        raise Exception(f"PDF '{nombre_doc}' resultó vacío.")
    return pdf_bytes


_MESES = ["enero","febrero","marzo","abril","mayo","junio","julio","agosto",
          "septiembre","octubre","noviembre","diciembre"]
_DIAS_NOMBRES = {
    'lunes': 'Lunes', 'martes': 'Martes', 'miercoles': 'Miércoles',
    'jueves': 'Jueves', 'viernes': 'Viernes', 'sabado': 'Sábado', 'domingo': 'Domingo',
}


def _ctx_contrato(contrato, es_plan_semilla: bool) -> dict:
    """Construye el contexto completo para el template contrato_trabajo.html."""
    empleado = contrato.empleado
    empresa = empleado.empresa
    hoy = datetime.date.today()
    fecha_espanol = f"{hoy.day:02d} de {_MESES[hoy.month - 1]} de {hoy.year}"
    comuna_emp = getattr(empresa, 'comuna', '') or getattr(empresa, 'ciudad', '') or ''
    ciudad = str(comuna_emp or getattr(empleado, 'comuna', '') or 'Santiago').strip().title()

    def fmt_fecha(fecha):
        if not fecha:
            return None
        return f"{fecha.day:02d} de {_MESES[fecha.month - 1]} de {fecha.year}"

    def fmt_pesos(valor):
        if not valor:
            return "$0"
        return f"${valor:,}".replace(",", ".")

    horario_formateado = []
    if contrato.distribucion_horario:
        for dia_key, datos in contrato.distribucion_horario.items():
            if datos.get('activo'):
                horario_formateado.append({
                    'dia_nombre': _DIAS_NOMBRES.get(dia_key, dia_key.title()),
                    'entrada': datos.get('entrada', ''),
                    'salida': datos.get('salida', ''),
                    'colacion': datos.get('colacion', 0),
                })

    return {
        'contrato': contrato,
        'empleado': empleado,
        'empresa': empresa,
        'es_plan_semilla': es_plan_semilla,
        'fecha_actual': fecha_espanol,
        'ciudad': ciudad,
        'tipo_contrato_texto': contrato.get_tipo_contrato_display(),
        'fecha_inicio_texto': fmt_fecha(contrato.fecha_inicio),
        'fecha_fin_texto': fmt_fecha(contrato.fecha_fin),
        'fecha_nacimiento_texto': fmt_fecha(empleado.fecha_nacimiento),
        'sueldo_base_texto': fmt_pesos(contrato.sueldo_base),
        'monto_quincena_texto': fmt_pesos(contrato.monto_quincena),
        'horario_formateado': horario_formateado,
    }


# ==========================================
# UTILIDADES DE VACACIONES (Art. 67-68 Código del Trabajo)
# ==========================================

# Feriados fijos chilenos (mes, día).
# No incluye Viernes/Sábado Santo (variables); se excluyen al calcular en el año real.
_FERIADOS_FIJOS_CL = frozenset([
    (1,  1),   # Año Nuevo
    (5,  1),   # Día del Trabajo
    (5,  21),  # Glorias Navales
    (6,  29),  # San Pedro y San Pablo
    (7,  16),  # Virgen del Carmen
    (8,  15),  # Asunción de la Virgen
    (9,  18),  # Fiestas Patrias
    (9,  19),  # Glorias del Ejército
    (10, 12),  # Encuentro de Dos Mundos
    (10, 31),  # Día de las Iglesias Evangélicas
    (11, 1),   # Todos los Santos
    (12, 8),   # Inmaculada Concepción
    (12, 25),  # Navidad
])


def _calcular_dias_habiles_vacacion(fecha_inicio, fecha_fin) -> int:
    """Días hábiles entre fecha_inicio y fecha_fin (inclusive) según ley chilena.

    Para vacaciones: hábil = cualquier día que NO sea domingo NI feriado fijo.
    Los sábados sí cuentan (Art. 67 Código del Trabajo).
    """
    dias = 0
    current = fecha_inicio
    delta_un_dia = datetime.timedelta(days=1)
    while current <= fecha_fin:
        if current.weekday() != 6 and (current.month, current.day) not in _FERIADOS_FIJOS_CL:
            dias += 1
        current += delta_un_dia
    return dias


def calcular_saldo_vacaciones(empleado) -> dict:
    """Saldo de vacaciones legales de un empleado (Art. 67-68 Código del Trabajo).

    Retorna:
        anos_servicio       — años completos desde fecha_ingreso
        dias_base           — 15 días × años_servicio
        dias_progresivos    — 1 día extra por cada 3 años sobre 10 (Art. 68)
        dias_devengados     — días_base + días_progresivos
        dias_usados         — suma de días_hábiles de registros APROBADO
        dias_disponibles    — devengados − usados (mínimo 0)
    """
    hoy = datetime.date.today()
    anos_servicio = (hoy - empleado.fecha_ingreso).days // 365

    dias_base = 15 * anos_servicio

    # Feriado progresivo: 1 día adicional por cada período completo de 3 años sobre 10
    dias_progresivos = max(0, (anos_servicio - 10) // 3) if anos_servicio >= 10 else 0

    dias_devengados = dias_base + dias_progresivos

    dias_usados = (
        VacacionEmpleado.objects
        .filter(
            empleado=empleado,
            estado='APROBADO',
            tipo__in=['VACACION_LEGAL', 'VACACION_PROGRESIVA'],
        )
        .aggregate(total=Sum('dias_habiles'))['total'] or 0
    )

    return {
        'anos_servicio':    anos_servicio,
        'dias_base':        dias_base,
        'dias_progresivos': dias_progresivos,
        'dias_devengados':  dias_devengados,
        'dias_usados':      int(dias_usados),
        'dias_disponibles': max(0, dias_devengados - int(dias_usados)),
    }


class DocumentoLegalViewSet(viewsets.ModelViewSet):
    queryset = DocumentoLegal.objects.all().order_by('-fecha_emision', '-creado_en')
    serializer_class = DocumentoLegalSerializer
    permission_classes = [IsAuthenticated]

    # Textos legales de cada causal para el PDF
    _CAUSAL_INFO = {
        '159_1':  ('Art. 159 N°1 — Mutuo acuerdo de las partes',
                   'Ambas partes acuerdan, de mutuo acuerdo, poner término al contrato de trabajo.', False),
        '159_2':  ('Art. 159 N°2 — Renuncia voluntaria del trabajador',
                   'El trabajador ha presentado su renuncia voluntaria al cargo.', False),
        '159_3':  ('Art. 159 N°3 — Muerte del trabajador',
                   'El contrato de trabajo termina por fallecimiento del trabajador.', False),
        '159_4':  ('Art. 159 N°4 — Vencimiento del plazo convenido',
                   'Ha vencido el plazo estipulado en el contrato de trabajo a plazo fijo.', False),
        '159_5':  ('Art. 159 N°5 — Conclusión del trabajo o servicio que dio origen al contrato',
                   'Ha concluido el trabajo, obra o servicio específico para el cual fue contratado el trabajador.', False),
        '159_6':  ('Art. 159 N°6 — Caso fortuito o fuerza mayor',
                   'Se ha producido un evento de caso fortuito o fuerza mayor que hace imposible continuar con la relación laboral.', False),
        '160_1a': ('Art. 160 N°1 a) — Falta de probidad del trabajador',
                   'El trabajador ha incurrido en conductas contrarias a la honradez e integridad que debe observar en el ejercicio de su cargo.', False),
        '160_1b': ('Art. 160 N°1 b) — Acoso sexual',
                   'El trabajador ha incurrido en conductas de acoso sexual, conforme a lo definido en el Artículo 2° del Código del Trabajo.', False),
        '160_1c': ('Art. 160 N°1 c) — Vías de hecho ejercidas por el trabajador',
                   'El trabajador ha ejercido vías de hecho en contra del empleador o de algún compañero de trabajo de la empresa.', False),
        '160_1d': ('Art. 160 N°1 d) — Injurias proferidas al empleador',
                   'El trabajador ha proferido injurias graves en contra del empleador, afectando su honor y dignidad.', False),
        '160_1e': ('Art. 160 N°1 e) — Conducta inmoral del trabajador',
                   'El trabajador ha incurrido en conductas inmorales graves que afectan a la empresa donde se desempeña.', False),
        '160_1f': ('Art. 160 N°1 f) — Conductas de acoso laboral',
                   'El trabajador ha incurrido en conductas de acoso laboral (mobbing), atentando contra la dignidad de otros trabajadores de la empresa.', False),
        '160_2':  ('Art. 160 N°2 — Negociaciones que ejecute el trabajador dentro del giro del negocio prohibidas por escrito',
                   'El trabajador ha realizado negociaciones dentro del giro del negocio de la empresa, en contravención a la prohibición expresa establecida en el contrato de trabajo.', False),
        '160_3':  ('Art. 160 N°3 — No concurrencia del trabajador a sus labores sin causa justificada',
                   'El trabajador no ha concurrido a sus labores sin causa justificada, configurándose la causal de inasistencias injustificadas establecida en el Código del Trabajo.', False),
        '160_4a': ('Art. 160 N°4 a) — Abandono del trabajo: salida intempestiva e injustificada',
                   'El trabajador ha abandonado el lugar de trabajo de forma intempestiva e injustificada durante la jornada laboral, sin permiso del empleador.', False),
        '160_4b': ('Art. 160 N°4 b) — Abandono del trabajo: negativa injustificada a trabajar',
                   'El trabajador se ha negado injustificadamente a realizar las faenas convenidas en el contrato de trabajo.', False),
        '160_5':  ('Art. 160 N°5 — Actos, omisiones o imprudencias temerarias que afecten la seguridad',
                   'El trabajador ha incurrido en actos, omisiones o imprudencias temerarias que afectan gravemente la seguridad o el funcionamiento del establecimiento, o la salud de los trabajadores.', False),
        '160_6':  ('Art. 160 N°6 — Perjuicio material causado intencionalmente',
                   'El trabajador ha causado intencionalmente perjuicio material en las instalaciones, maquinarias, herramientas, útiles de trabajo, productos o mercaderías de la empresa.', False),
        '160_7':  ('Art. 160 N°7 — Incumplimiento grave de las obligaciones que impone el contrato',
                   'El trabajador ha incurrido en incumplimiento grave de las obligaciones que le impone el contrato de trabajo.', False),
        '161_1':  ('Art. 161 inciso 1° — Necesidades de la empresa, establecimiento o servicio',
                   'La empresa, por razones derivadas de la racionalización o modernización de la misma, bajas en la productividad, cambios en las condiciones del mercado o de la economía, o que hagan necesaria la separación de uno o más trabajadores, ha decidido poner término al contrato de trabajo.', True),
        '161_2':  ('Art. 161 inciso 2° — Desahucio del empleador',
                   'El empleador, en ejercicio de la facultad contemplada en el inciso segundo del Artículo 161 del Código del Trabajo, pone término al contrato de trabajo mediante desahucio.', True),
        '163bis': ('Art. 163 bis — Liquidación concursal del empleador',
                   'La empresa ha sido sometida a un procedimiento concursal de liquidación de sus bienes por resolución judicial, lo que determina el término del contrato de trabajo conforme a lo dispuesto en el Artículo 163 bis del Código del Trabajo.', True),
    }

    def get_queryset(self):
        # Solo documentos de empleados que pertenecen al usuario autenticado
        queryset = DocumentoLegal.objects.filter(
            empleado__empresa__owner=self.request.user
        ).order_by('-fecha_emision', '-creado_en')
        empleado_id = self.request.query_params.get('empleado', None)
        if empleado_id is not None:
            queryset = queryset.filter(empleado_id=empleado_id)
        return queryset

    @action(detail=True, methods=['get'])
    def generar_pdf(self, request, pk=None):
        try:
            documento = self.get_object()
            empleado = documento.empleado
            empresa = empleado.empresa

            meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            hoy = documento.fecha_emision
            fecha_espanol = f"{hoy.day:02d} de {meses[hoy.month - 1]} de {hoy.year}"

            comuna_emp = getattr(empresa, 'comuna', '') or getattr(empresa, 'ciudad', '') or ''
            comuna_empl = getattr(empleado, 'comuna', '') or ''
            ciudad_segura = str(comuna_emp or comuna_empl or 'Santiago').strip().title()
            es_plan_semilla = _es_plan_semilla(request.user)

            context = {
                'documento': documento,
                'empleado': empleado,
                'empresa': empresa,
                'fecha_actual': fecha_espanol,
                'ciudad': ciudad_segura,
                'es_plan_semilla': es_plan_semilla,
            }

            if documento.tipo == 'DESPIDO':
                # Contexto enriquecido para carta_despido.html
                codigo = documento.causal_articulo or ''
                causal_label, causal_descripcion, requiere_indemnizacion = self._CAUSAL_INFO.get(
                    codigo, (documento.causal_legal or '—', '', False)
                )
                def _fmt_pesos(v):
                    if not v:
                        return '$ 0'
                    return f'$ {v:,}'.replace(',', '.')
                monto_anos  = documento.monto_indemnizacion_anos or 0
                monto_sust  = documento.monto_indemnizacion_sustitutiva or 0
                # Cargo desde contrato si existe
                try:
                    contrato_cargo = documento.empleado.contrato.cargo
                except Exception:
                    contrato_cargo = None
                # Fecha último día en español
                fecha_ultimo_dia_texto = None
                if documento.fecha_ultimo_dia:
                    f = documento.fecha_ultimo_dia
                    fecha_ultimo_dia_texto = f"{f.day:02d} de {meses[f.month - 1]} de {f.year}"
                context.update({
                    'causal_label': causal_label,
                    'causal_descripcion': causal_descripcion,
                    'requiere_indemnizacion': requiere_indemnizacion,
                    'monto_anos_texto': _fmt_pesos(monto_anos),
                    'monto_sustitutiva_texto': _fmt_pesos(monto_sust),
                    'monto_total_texto': _fmt_pesos(monto_anos + monto_sust),
                    'fecha_ultimo_dia_texto': fecha_ultimo_dia_texto,
                    'contrato_cargo': contrato_cargo,
                })
                template = get_template('carta_despido.html')
            else:
                template = get_template('documento_legal.html')

            html = template.render(context)

            response = HttpResponse(content_type='application/pdf')
            nombre_archivo = f'{documento.tipo}_{empleado.rut}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            pisa_status = pisa.CreatePDF(html, dest=response)

            if pisa_status.err:
                return HttpResponse('Error al generar el PDF.', status=500)
            
            return response

        except Exception as e:
            return Response({'error': f'Error al generar PDF: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ==========================================
# VACACIONES Y PERMISOS
# ==========================================
class VacacionViewSet(viewsets.ModelViewSet):
    serializer_class = None  # se asigna abajo tras importar el serializer
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        from .serializers import VacacionSerializer
        return VacacionSerializer

    def get_queryset(self):
        qs = VacacionEmpleado.objects.filter(
            empresa__owner=self.request.user
        ).order_by('-fecha_inicio')
        empleado_id = self.request.query_params.get('empleado')
        if empleado_id:
            qs = qs.filter(empleado_id=empleado_id)
        return qs

    def create(self, request, *args, **kwargs):
        if not _plan_permite(request.user, 2):
            return Response(
                {'error': 'La gestión de vacaciones y permisos está disponible desde el plan Starter. Mejora tu suscripción para acceder.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        """Auto-calcula días hábiles si el cliente no los envía."""
        fecha_inicio = serializer.validated_data.get('fecha_inicio')
        fecha_fin    = serializer.validated_data.get('fecha_fin')
        dias = serializer.validated_data.get('dias_habiles') or 0
        if fecha_inicio and fecha_fin and not dias:
            dias = _calcular_dias_habiles_vacacion(fecha_inicio, fecha_fin)
        serializer.save(dias_habiles=dias)

    @action(detail=False, methods=['get'], url_path='saldo')
    def saldo(self, request):
        """GET /api/vacaciones/saldo/?empleado=<id>
        Retorna el saldo de vacaciones del empleado.
        """
        if not _plan_permite(request.user, 2):
            return Response(
                {'error': 'La gestión de vacaciones está disponible desde el plan Starter.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        empleado_id = request.query_params.get('empleado')
        if not empleado_id:
            return Response({'error': 'Parámetro empleado requerido.'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            empleado = Empleado.objects.get(pk=empleado_id, empresa__owner=request.user)
        except Empleado.DoesNotExist:
            return Response({'error': 'Empleado no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(calcular_saldo_vacaciones(empleado))

    @action(detail=True, methods=['get'], url_path='generar_pdf')
    def generar_pdf(self, request, pk=None):
        """GET /api/vacaciones/<id>/generar_pdf/
        Genera y devuelve el comprobante de vacaciones en PDF.
        """
        try:
            vacacion = self.get_object()
            empleado = vacacion.empleado
            empresa  = vacacion.empresa
            es_semilla = _es_plan_semilla(request.user)

            hoy = datetime.date.today()
            fecha_hoy_texto = f"{hoy.day:02d} de {_MESES[hoy.month - 1]} de {hoy.year}"

            def _fmt_fecha(f):
                if not f:
                    return '—'
                return f"{f.day:02d} de {_MESES[f.month - 1]} de {f.year}"

            context = {
                'vacacion':          vacacion,
                'empleado':          empleado,
                'empresa':           empresa,
                'fecha_actual':      fecha_hoy_texto,
                'fecha_inicio_texto': _fmt_fecha(vacacion.fecha_inicio),
                'fecha_fin_texto':    _fmt_fecha(vacacion.fecha_fin),
                'ciudad': str(
                    getattr(empresa, 'ciudad', '') or
                    getattr(empresa, 'comuna', '') or
                    getattr(empleado, 'comuna', '') or 'Santiago'
                ).strip().title(),
                'es_plan_semilla': es_semilla,
            }

            template = get_template('comprobante_vacaciones.html')
            html = template.render(context)

            response = HttpResponse(content_type='application/pdf')
            nombre = f'vacacion_{empleado.rut}_{vacacion.fecha_inicio}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{nombre}"'

            pisa_status = pisa.CreatePDF(html, dest=response)
            if pisa_status.err:
                return HttpResponse('Error al generar el PDF.', status=500)

            return response

        except Exception as e:
            return Response(
                {'error': f'Error al generar PDF: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class EmpresaViewSet(viewsets.ModelViewSet):
    serializer_class = EmpresaSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.request.query_params.get('incluir_inactivas') == 'true':
            return Empresa.objects.filter(owner=self.request.user)
        return Empresa.objects.filter(owner=self.request.user, activo=True)
    
    @action(detail=True, methods=['post'])
    def reactivar(self, request, pk=None):
        try:
            empresa = Empresa.objects.get(pk=pk, owner=request.user)
            empresa.activo = True
            empresa.save()
            return Response({"mensaje": "Empresa reactivada correctamente"}, status=status.HTTP_200_OK)
        except Empresa.DoesNotExist:
            return Response({"error": "Empresa no encontrada"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['patch'], url_path='configurar-firma')
    def configurar_firma(self, request, pk=None):
        """Guarda la firma dibujada del representante legal de la empresa."""
        empresa = self.get_object()

        firma_imagen = request.data.get('firma_imagen', '').strip()
        nombre       = request.data.get('firma_firmante_nombre', '').strip()
        cargo        = request.data.get('firma_firmante_cargo', '').strip()

        if not firma_imagen:
            return Response({'error': 'La imagen de firma es requerida.'}, status=status.HTTP_400_BAD_REQUEST)
        if not firma_imagen.startswith('data:image/'):
            return Response({'error': 'Formato de imagen inválido.'}, status=status.HTTP_400_BAD_REQUEST)
        if len(firma_imagen) > 500_000:
            return Response({'error': 'La imagen de firma es demasiado grande.'}, status=status.HTTP_400_BAD_REQUEST)

        empresa.firma_imagen          = firma_imagen
        empresa.firma_firmante_nombre = nombre
        empresa.firma_firmante_cargo  = cargo
        empresa.firma_configurada_en  = timezone.now()
        empresa.save(update_fields=['firma_imagen', 'firma_firmante_nombre',
                                    'firma_firmante_cargo', 'firma_configurada_en'])

        serializer = self.get_serializer(empresa)
        return Response(serializer.data)

    def perform_create(self, serializer):
        cliente = getattr(self.request.user, 'perfil_cliente', None)
        
        # 1. REGLA DE NEGOCIO: Límite de empresas según el Plan
        if cliente and cliente.plan:
            total_empresas = Empresa.objects.filter(owner=self.request.user).count()
            
            # Definir límite según el ID del plan (1: Semilla, 2: Pyme, 3: Corporativo)
            limite_empresas = cliente.plan.max_empresas
                        
            if total_empresas >= limite_empresas:
                raise ValidationError({'error': f'Tu {cliente.plan.nombre} permite administrar un máximo de {limite_empresas} empresas. Actualiza tu plan para registrar más.'})

        # 2. Convertir a mayúsculas
        datos_mayusculas = {k: (v.upper() if isinstance(v, str) else v) for k, v in serializer.validated_data.items()}
        rut_raw = self.request.data.get('rut', '')
        
        # 3. REGLA DE NEGOCIO: No repetir RUT en el mismo panel
        if rut_raw:
            rut_form = formatear_rut(rut_raw)
            if Empresa.objects.filter(rut=rut_form).exists():
                raise ValidationError({'error': 'Ya existe una empresa registrada con este RUT en el sistema. Contacta a soporte si crees que esto es un error.'})
            datos_mayusculas['rut'] = rut_form
            
        serializer.save(owner=self.request.user, **datos_mayusculas)

    # SOFT-DELETE: En vez de eliminar la empresa, la marcamos como inactiva
    def destroy(self, request, *args, **kwargs):
        empresa = self.get_object()
        empresa.activo = False
        empresa.save()
        return Response({"mensaje": "Empresa desactivada correctamente"}, status=status.HTTP_200_OK)
            
    def perform_update(self, serializer):
        datos_mayusculas = {k: (v.upper() if isinstance(v, str) else v) for k, v in serializer.validated_data.items()}
        rut_raw = self.request.data.get('rut', '')
        
        if rut_raw:
            rut_form = formatear_rut(rut_raw)
            if Empresa.objects.filter(owner=self.request.user, rut=rut_form).exclude(id=serializer.instance.id).exists():
                raise ValidationError({'error': 'Ya tienes otra empresa registrada con este RUT.'})
            datos_mayusculas['rut'] = rut_form
            
        serializer.save(**datos_mayusculas)

class EmpleadoViewSet(viewsets.ModelViewSet):
    serializer_class = EmpleadoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        rechazos_qs = SolicitudFirma.objects.filter(
            empleado=OuterRef('pk'),
            estado='RECHAZADO',
        )
        return Empleado.objects.filter(empresa__owner=self.request.user).annotate(
            tiene_rechazos_pendientes=Exists(rechazos_qs)
        )

    def perform_create(self, serializer):
        datos_mayusculas = {k: (v.upper() if isinstance(v, str) else v) for k, v in serializer.validated_data.items()}

        empresa_destino = serializer.validated_data.get('empresa')

        rut_raw = self.request.data.get('rut', '')
        if rut_raw:
            if not validar_rut(rut_raw):
                raise ValidationError({'error': 'El RUT ingresado no es válido.'})
            rut_form = formatear_rut(rut_raw)

            if Empleado.objects.filter(empresa=empresa_destino, rut=rut_form).exists():
                raise ValidationError({'error': 'Este trabajador ya está registrado en esta empresa.'})

            datos_mayusculas['rut'] = rut_form

        telefono = datos_mayusculas.get('numero_telefono')
        if telefono and isinstance(telefono, str):
            solo_digitos = re.sub(r'[^0-9]', '', telefono)
            datos_mayusculas['numero_telefono'] = f'+56{solo_digitos[-9:]}' if solo_digitos else None

        serializer.save(**datos_mayusculas)

    def perform_update(self, serializer):
        datos_mayusculas = {k: (v.upper() if isinstance(v, str) else v) for k, v in serializer.validated_data.items()}

        empresa_destino = serializer.validated_data.get('empresa', serializer.instance.empresa)

        rut_raw = self.request.data.get('rut', '')
        if rut_raw:
            if not validar_rut(rut_raw):
                raise ValidationError({'error': 'El RUT ingresado no es válido.'})
            rut_form = formatear_rut(rut_raw)

            if Empleado.objects.filter(empresa=empresa_destino, rut=rut_form).exclude(id=serializer.instance.id).exists():
                raise ValidationError({'error': 'Ya existe otro trabajador con este RUT en esta empresa.'})

            datos_mayusculas['rut'] = rut_form

        # Normalizar teléfono: extraer solo los últimos 9 dígitos y anteponer +56
        telefono = datos_mayusculas.get('numero_telefono')
        if telefono and isinstance(telefono, str):
            solo_digitos = re.sub(r'[^0-9]', '', telefono)
            datos_mayusculas['numero_telefono'] = f'+56{solo_digitos[-9:]}' if solo_digitos else None

        serializer.save(**datos_mayusculas)

    @action(detail=False, methods=['post'])
    def carga_masiva(self, request):
        if not _plan_permite(request.user, 3):
            return Response(
                {'error': 'La importación masiva por Excel está disponible desde el plan Pyme. Mejora tu suscripción para acceder.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        try:
            data = request.data[0] if isinstance(request.data, list) else request.data
            empresa_id = data.get('empresa')
            archivo_excel = request.FILES.get('file') or data.get('file')

            if not archivo_excel or not empresa_id:
                return Response({'error': 'Falta el archivo o la empresa.'}, status=400)

            MAX_EXCEL_MB = 5
            if hasattr(archivo_excel, 'size') and archivo_excel.size > MAX_EXCEL_MB * 1024 * 1024:
                return Response({'error': f'El archivo no puede superar {MAX_EXCEL_MB} MB.'}, status=400)

            MIME_EXCEL = {
                'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'application/vnd.ms-excel',
            }
            if hasattr(archivo_excel, 'content_type') and archivo_excel.content_type not in MIME_EXCEL:
                return Response({'error': 'Solo se aceptan archivos Excel (.xlsx o .xls).'}, status=400)

            empresa = Empresa.objects.get(id=empresa_id, owner=request.user)
            cliente = getattr(request.user, 'perfil_cliente', None)
            limite_trabajadores = cliente.plan.limite_trabajadores if (cliente and cliente.plan) else 1000

            # Leemos el Excel
            df = pd.read_excel(archivo_excel).fillna('')
            registros = df.to_dict('records')

            MAX_FILAS = 500
            if len(registros) > MAX_FILAS:
                return Response({'error': f'El archivo no puede tener más de {MAX_FILAS} filas por importación.'}, status=400)
            
            empleados_creados = 0
            empleados_actualizados = 0
            limite_alcanzado = False
            errores = []
            
            # Preparamos el mapa de empleados actuales para evitar duplicados
            empleados_bd = Empleado.objects.filter(empresa=empresa)
            mapa_empleados = { limpiar_rut(emp.rut): emp for emp in empleados_bd }
            
            siguiente_ficha = (empleados_bd.aggregate(Max('ficha_numero'))['ficha_numero__max'] or 0) + 1

            with transaction.atomic():
                total_actual = empleados_bd.count()

                for fila_num, row in enumerate(registros, start=2):  # start=2 porque fila 1 es el header del Excel
                    # --- NORMALIZADOR DE COLUMNAS ---
                    # Creamos un nuevo diccionario con todas las llaves en minúsculas y sin espacios
                    # Así row.get('email') funcionará aunque el Excel diga "Email", " EMAIL" o "email"
                    row_norm = { str(k).strip().lower().replace(' ', '_'): v for k, v in row.items() }
                    
                    rut_raw = str(row_norm.get('rut', '')).strip()
                    if not rut_raw: continue
                    
                    if not validar_rut(rut_raw):
                        continue # Saltamos RUTs inválidos

                    rut_limpio = limpiar_rut(rut_raw)
                    rut_formateado = formatear_rut(rut_raw)

                    # Extraer datos usando las llaves normalizadas
                    nombres = str(row_norm.get('nombres', '')).strip().upper()
                    ap_paterno = str(row_norm.get('apellido_paterno', '')).strip().upper()
                    email_dato = str(row_norm.get('email', '')).strip().lower() # <--- AQUÍ SE CORRIGE EL EMAIL

                    # Lógica de tipos de datos
                    try:
                        sueldo = int(float(row_norm.get('sueldo_base', 0)))
                        if sueldo < 0:
                            errores.append(f"Fila {fila_num}: sueldo_base no puede ser negativo.")
                            continue
                        horas_raw = row_norm.get('horas_laborales', 44)
                        horas = int(horas_raw)
                        if horas <= 0 or horas > 168:
                            errores.append(f"Fila {fila_num}: horas_laborales debe estar entre 1 y 168.")
                            continue
                    except (ValueError, TypeError):
                        errores.append(f"Fila {fila_num}: sueldo_base u horas_laborales contienen valores no numéricos.")
                        continue

                    raw_cuenta = row_norm.get('numero_cuenta', '')
                    try:
                        # Si es numérico (ej. 12345.0), lo pasamos a float, luego a entero (quita el .0) y luego a texto
                        num_cuenta = str(int(float(raw_cuenta)))
                    except (ValueError, TypeError):
                        # Si tiene letras o guiones (ej. "Chequera-123") o está vacío, lo dejamos como texto normal
                        num_cuenta = str(raw_cuenta).strip()

                    # AQUÍ SE ASIGNAN TODOS LOS CAMPOS QUE ESTABAN EN NULL
                    nuevos_datos = {
                        'nombres': nombres,
                        'apellido_paterno': ap_paterno,
                        'apellido_materno': str(row_norm.get('apellido_materno', '')).strip().upper(),
                        'email': email_dato,
                        'sexo': str(row_norm.get('sexo', 'M')).strip().upper()[:1],
                        'nacionalidad': str(row_norm.get('nacionalidad', 'CHILENA')).strip().upper(),
                        'fecha_nacimiento': estandarizar_fecha(row_norm.get('fecha_nacimiento')),
                        'fecha_ingreso': estandarizar_fecha(row_norm.get('fecha_ingreso')) or datetime.date.today(),
                        'departamento': str(row_norm.get('departamento', '')).strip().upper(),
                        'sucursal': str(row_norm.get('sucursal', '')).strip().upper(),
                        'cargo': str(row_norm.get('cargo', '')).strip().upper(),
                        'sueldo_base': sueldo,
                        'horas_laborales': horas,
                        'forma_pago': str(row_norm.get('forma_pago', 'TRANSFERENCIA')).strip().upper(),
                        'banco': str(row_norm.get('banco', '')).strip().upper(),
                        'tipo_cuenta': str(row_norm.get('tipo_cuenta', '')).strip().upper(),
                        'numero_cuenta': num_cuenta
                    }

                    empleado_existente = mapa_empleados.get(rut_limpio)

                    if empleado_existente:
                        # ACTUALIZAR
                        for key, value in nuevos_datos.items():
                            setattr(empleado_existente, key, value)
                        empleado_existente.save()
                        empleados_actualizados += 1
                    else:
                        # CREAR (Validando límite de plan)
                        if total_actual >= limite_trabajadores:
                            limite_alcanzado = True
                            continue
                        
                        Empleado.objects.create(
                            rut=rut_formateado,
                            empresa=empresa,
                            ficha_numero=siguiente_ficha,
                            **nuevos_datos
                        )
                        siguiente_ficha += 1
                        empleados_creados += 1
                        total_actual += 1

            return Response({
                'agregados': empleados_creados,
                'actualizados': empleados_actualizados,
                'limite_alcanzado': limite_alcanzado,
                'errores': errores,
            }, status=200)

        except Exception:
            return Response({'error': 'Error procesando el archivo. Revisa el formato e inténtalo de nuevo.'}, status=500)
                
   # ====================================================
    # DISPONIBILIDAD DE DOCUMENTOS POR TRABAJADOR
    # ====================================================
    @action(detail=True, methods=['get'])
    def documentos_disponibles(self, request, pk=None):
        empleado = self.get_object()
        contrato = Contrato.objects.filter(empleado=empleado).first()

        data = {
            'tiene_contrato': contrato is not None,
            'tiene_anexo_40h': bool(contrato and contrato.archivo_anexo_40h),
            'cantidad_liquidaciones': Liquidacion.objects.filter(empleado=empleado).count(),
            'cantidad_amonestaciones': DocumentoLegal.objects.filter(empleado=empleado, tipo='AMONESTACION').count(),
            'tiene_despido': DocumentoLegal.objects.filter(empleado=empleado, tipo='DESPIDO').exists(),
            'tiene_mutuo_acuerdo': DocumentoLegal.objects.filter(empleado=empleado, tipo='MUTUO_ACUERDO').exists(),
            'cantidad_constancias': DocumentoLegal.objects.filter(empleado=empleado, tipo='CONSTANCIA').count(),
            'cantidad_anexos_contrato': AnexoContrato.objects.filter(contrato=contrato).count() if contrato else 0,
        }
        return Response(data)

    @action(detail=True, methods=['get'], url_path='historial_salarial')
    def historial_salarial(self, request, pk=None):
        empleado = self.get_object()
        contrato = Contrato.objects.filter(empleado=empleado).first()

        liqs = list(
            Liquidacion.objects.filter(empleado=empleado)
            .order_by('anio', 'mes')
            .values('mes', 'anio', 'sueldo_base', 'total_haberes', 'total_descuentos',
                    'sueldo_liquido', 'dias_trabajados')
        )

        periodos = []
        prev_liq = None
        for liq in liqs:
            delta_pct = None
            if prev_liq is not None and prev_liq > 0:
                delta_pct = round((liq['sueldo_liquido'] - prev_liq) / prev_liq * 100, 1)
            periodos.append({**liq, 'delta_pct': delta_pct})
            prev_liq = liq['sueldo_liquido']

        liquidos = [p['sueldo_liquido'] for p in periodos]
        promedio = round(sum(liquidos) / len(liquidos)) if liquidos else 0

        # Tendencia: promedio últimos 3 meses vs los 3 anteriores
        tendencia_3m = None
        if len(liquidos) >= 6:
            avg_rec = sum(liquidos[-3:]) / 3
            avg_ant = sum(liquidos[-6:-3]) / 3
            if avg_ant > 0:
                tendencia_3m = round((avg_rec - avg_ant) / avg_ant * 100, 1)
        elif len(liquidos) >= 2:
            tendencia_3m = periodos[-1]['delta_pct']

        return Response({
            'contrato_sueldo_base': contrato.sueldo_base if contrato else None,
            'contrato_tipo': contrato.tipo_contrato if contrato else None,
            'promedio_liquido': promedio,
            'tendencia_3m': tendencia_3m,
            'periodos': periodos,
        })

   # ====================================================
    # MOTOR DOCUMENTAL PERSISTENTE (CON PLANTILLAS REALES)
    # ====================================================

    def _html_a_pdf(self, html_string, nombre_doc):
        """Convierte HTML a bytes PDF con xhtml2pdf. Lanza excepción clara si falla."""
        resultado = io.BytesIO()
        status = pisa.pisaDocument(io.BytesIO(html_string.encode("UTF-8")), resultado)
        if status.err:
            raise Exception(
                f"xhtml2pdf reportó {status.err} error(s) generando '{nombre_doc}'. "
                f"Revisar el template y el contexto pasado."
            )
        pdf_bytes = resultado.getvalue()
        if not pdf_bytes:
            raise Exception(
                f"PDF '{nombre_doc}' resultó vacío tras la conversión. "
                f"xhtml2pdf no reportó error pero el resultado es b\"\". "
                f"Posible problema de encoding o template vacío."
            )
        return pdf_bytes

    def _obtener_o_generar_documento(self, empleado, tipo_documento, user=None):
        """Revisa si el PDF ya existe en la BD. Si no, lo genera usando los templates HTML reales."""

        empresa = empleado.empresa
        es_plan_semilla = _es_plan_semilla(user) if user else False

        # --- LÓGICA PARA CONTRATOS ---
        if tipo_documento == 'contrato':
            try:
                contrato = Contrato.objects.get(empleado=empleado)
            except Contrato.DoesNotExist:
                raise Exception(f"El trabajador {empleado.nombres} no tiene contrato registrado.")
            if contrato.archivo_contrato:
                try:
                    return contrato.archivo_contrato.read()
                except Exception:
                    pass

            context = _ctx_contrato(contrato, es_plan_semilla)
            html_string = render_to_string('contrato_trabajo.html', context)
            pdf_bytes = self._html_a_pdf(html_string, f'Contrato_{empleado.rut}')
            contrato.archivo_contrato.save(f"Contrato_{empleado.rut}.pdf", ContentFile(pdf_bytes))
            return pdf_bytes

        # --- LÓGICA PARA ANEXOS 40 HORAS ---
        elif tipo_documento == 'anexo_40h':
            try:
                contrato = Contrato.objects.get(empleado=empleado)
            except Contrato.DoesNotExist:
                raise Exception(f"El trabajador {empleado.nombres} no tiene contrato registrado.")
            if contrato.archivo_anexo_40h:
                try:
                    return contrato.archivo_anexo_40h.read()
                except Exception:
                    pass

            context = _ctx_contrato(contrato, es_plan_semilla)
            html_string = render_to_string('anexo_40h.html', context)
            pdf_bytes = self._html_a_pdf(html_string, f'Anexo_40h_{empleado.rut}')
            contrato.archivo_anexo_40h.save(f"Anexo_40h_{empleado.rut}.pdf", ContentFile(pdf_bytes))
            return pdf_bytes

        # --- LÓGICA PARA LIQUIDACIONES (MES ACTUAL) ---
        elif tipo_documento == 'liquidacion_actual':
            hoy = datetime.date.today()
            try:
                liquidacion = Liquidacion.objects.get(empleado=empleado, mes=hoy.month, anio=hoy.year)
            except Liquidacion.DoesNotExist:
                raise Exception(f"No existe liquidación del mes actual para {empleado.nombres}.")
            if liquidacion.archivo_pdf:
                try:
                    return liquidacion.archivo_pdf.read()
                except Exception:
                    pass

            try:
                liquido_palabras = num2words(liquidacion.sueldo_liquido, lang='es')
            except Exception:
                liquido_palabras = str(liquidacion.sueldo_liquido)
            contrato_liq = Contrato.objects.filter(empleado=empleado).first()
            meses_liq = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
            det_no_imp = liquidacion.detalle_haberes_no_imponibles
            if not isinstance(det_no_imp, list): det_no_imp = []
            det_otros = liquidacion.detalle_otros_descuentos
            if not isinstance(det_otros, list): det_otros = []
            context = {
                'empleado': empleado, 'empresa': empresa,
                'liquidacion': liquidacion, 'contrato': contrato_liq,
                'mes_nombre': meses_liq[liquidacion.mes - 1].upper(),
                'liquido_palabras': liquido_palabras,
                'total_no_imponible': sum(int(i.get('valor', 0)) for i in det_no_imp if isinstance(i, dict)),
                'total_ley': (liquidacion.afp_monto or 0) + (liquidacion.salud_monto or 0) + (liquidacion.seguro_cesantia or 0) + (liquidacion.impuesto_unico or 0),
                'total_otros_dsctos': (liquidacion.anticipo_quincena or 0) + sum(int(i.get('valor', 0)) for i in det_otros if isinstance(i, dict)),
                'es_plan_semilla': es_plan_semilla,
            }
            html_string = render_to_string('liquidacion.html', context)
            pdf_bytes = self._html_a_pdf(html_string, f'Liquidacion_{hoy.month}_{hoy.year}_{empleado.rut}')
            liquidacion.archivo_pdf.save(
                f"Liquidacion_{hoy.month}_{hoy.year}_{empleado.rut}.pdf", ContentFile(pdf_bytes)
            )
            return pdf_bytes

        # --- LÓGICA PARA LIQUIDACIONES HISTÓRICAS ---
        elif tipo_documento.startswith('liquidacion_historica_'):
            _, _, mes_str, anio_str = tipo_documento.split('_')
            mes_hist, anio_hist = int(mes_str), int(anio_str)

            try:
                liquidacion = Liquidacion.objects.get(empleado=empleado, mes=mes_hist, anio=anio_hist)
            except Liquidacion.DoesNotExist:
                raise Exception(f"No existe liquidación {mes_hist}/{anio_hist} para {empleado.nombres}.")
            if liquidacion.archivo_pdf:
                try:
                    return liquidacion.archivo_pdf.read()
                except Exception:
                    pass

            try:
                liquido_palabras = num2words(liquidacion.sueldo_liquido, lang='es')
            except Exception:
                liquido_palabras = str(liquidacion.sueldo_liquido)
            contrato_liq = Contrato.objects.filter(empleado=empleado).first()
            meses_liq = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
            det_no_imp = liquidacion.detalle_haberes_no_imponibles
            if not isinstance(det_no_imp, list): det_no_imp = []
            det_otros = liquidacion.detalle_otros_descuentos
            if not isinstance(det_otros, list): det_otros = []
            context = {
                'empleado': empleado, 'empresa': empresa,
                'liquidacion': liquidacion, 'contrato': contrato_liq,
                'mes_nombre': meses_liq[liquidacion.mes - 1].upper(),
                'liquido_palabras': liquido_palabras,
                'total_no_imponible': sum(int(i.get('valor', 0)) for i in det_no_imp if isinstance(i, dict)),
                'total_ley': (liquidacion.afp_monto or 0) + (liquidacion.salud_monto or 0) + (liquidacion.seguro_cesantia or 0) + (liquidacion.impuesto_unico or 0),
                'total_otros_dsctos': (liquidacion.anticipo_quincena or 0) + sum(int(i.get('valor', 0)) for i in det_otros if isinstance(i, dict)),
                'es_plan_semilla': es_plan_semilla,
            }
            html_string = render_to_string('liquidacion.html', context)
            pdf_bytes = self._html_a_pdf(html_string, f'Liquidacion_{mes_hist}_{anio_hist}_{empleado.rut}')
            liquidacion.archivo_pdf.save(
                f"Liquidacion_{mes_hist}_{anio_hist}_{empleado.rut}.pdf", ContentFile(pdf_bytes)
            )
            return pdf_bytes

        # --- LÓGICA PARA CARTAS DE AMONESTACIÓN ---
        elif tipo_documento == 'amonestacion':
            doc_legal = DocumentoLegal.objects.filter(
                empleado=empleado, tipo='AMONESTACION'
            ).order_by('-fecha_emision').first()
            if doc_legal is None:
                raise Exception(f"El trabajador {empleado.nombres} no tiene amonestaciones registradas.")
            if doc_legal.archivo_pdf:
                try:
                    return doc_legal.archivo_pdf.read()
                except Exception:
                    pass

            pdf_bytes = self._pdf_para_documento_legal(doc_legal, es_plan_semilla)
            doc_legal.archivo_pdf.save(f"Amonestacion_{empleado.rut}.pdf", ContentFile(pdf_bytes))
            return pdf_bytes

        raise Exception(f"Tipo de documento no soportado: '{tipo_documento}'")
    
    # ====================================================
    # HELPERS PDF PARA DOCUMENTOS LEGALES Y ANEXOS
    # ====================================================
    def _pdf_para_documento_legal(self, doc, es_plan_semilla):
        empleado = doc.empleado
        empresa = empleado.empresa
        meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        hoy = doc.fecha_emision
        fecha_espanol = f"{hoy.day:02d} de {meses[hoy.month - 1]} de {hoy.year}"
        ciudad = str(getattr(empresa, 'comuna', '') or getattr(empresa, 'ciudad', '') or 'Santiago').strip().title()
        context = {
            'documento': doc, 'empleado': empleado, 'empresa': empresa,
            'fecha_actual': fecha_espanol, 'ciudad': ciudad,
            'es_plan_semilla': es_plan_semilla,
        }
        if doc.tipo == 'DESPIDO':
            codigo = doc.causal_articulo or ''
            causal_label, causal_descripcion, requiere_indemnizacion = self._CAUSAL_INFO.get(
                codigo, (doc.causal_legal or '—', '', False)
            )
            def _fmt(v):
                return f'$ {v:,}'.replace(',', '.') if v else '$ 0'
            fecha_ult = None
            if doc.fecha_ultimo_dia:
                f = doc.fecha_ultimo_dia
                fecha_ult = f"{f.day:02d} de {meses[f.month - 1]} de {f.year}"
            try:
                contrato_cargo = doc.empleado.contrato.cargo
            except Exception:
                contrato_cargo = None
            context.update({
                'causal_label': causal_label,
                'causal_descripcion': causal_descripcion,
                'requiere_indemnizacion': requiere_indemnizacion,
                'monto_anos_texto': _fmt(doc.monto_indemnizacion_anos or 0),
                'monto_sustitutiva_texto': _fmt(doc.monto_indemnizacion_sustitutiva or 0),
                'monto_total_texto': _fmt((doc.monto_indemnizacion_anos or 0) + (doc.monto_indemnizacion_sustitutiva or 0)),
                'fecha_ultimo_dia_texto': fecha_ult,
                'contrato_cargo': contrato_cargo,
            })
            html = render_to_string('carta_despido.html', context)
        else:
            html = render_to_string('documento_legal.html', context)
        return self._html_a_pdf(html, f'{doc.tipo}_{empleado.rut}_{doc.fecha_emision}')

    def _pdf_para_anexo_contrato(self, anexo, es_plan_semilla):
        contrato = anexo.contrato
        empleado = contrato.empleado
        empresa = empleado.empresa
        meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
        hoy = anexo.fecha_emision
        fecha_espanol = f"{hoy.day:02d} de {meses[hoy.month - 1]} de {hoy.year}"
        ciudad = str(getattr(empresa, 'comuna', '') or getattr(empresa, 'ciudad', '') or 'Santiago').strip().title()
        context = {
            'anexo': anexo, 'contrato': contrato, 'empleado': empleado, 'empresa': empresa,
            'fecha_actual': fecha_espanol, 'ciudad': ciudad,
            'es_plan_semilla': es_plan_semilla,
        }
        html = render_to_string('anexo_contrato.html', context)
        return self._html_a_pdf(html, f'AnexoContrato_{empleado.rut}_{hoy}')

    # ====================================================
    # ENDPOINT: DESCARGA MASIVA Y EXPEDIENTES (ZIP)
    # ====================================================
    @action(detail=False, methods=['post'])
    def descarga_masiva(self, request):
        """
        POST body:
          empleados: [id, ...]
          empresa_id: int
          documentos: lista de tipos a incluir:
            'contrato', 'anexo_40h', 'liquidaciones',
            'amonestaciones', 'despidos', 'mutuo_acuerdo',
            'constancias', 'anexos_contrato'
          cantidad_liquidaciones: int (cuántas liquidaciones recientes incluir)
        """
        try:
            empleados_ids = request.data.get('empleados', [])
            empresa_id = request.data.get('empresa_id')
            documentos = request.data.get('documentos', [])
            cantidad_liquidaciones = int(request.data.get('cantidad_liquidaciones', 1))

            if not empleados_ids or not empresa_id:
                return Response({'error': 'Faltan IDs de trabajadores o empresa'}, status=400)
            if not documentos:
                return Response({'error': 'Debes seleccionar al menos un tipo de documento'}, status=400)

            MAX_EMPLEADOS_ZIP = 50
            if len(empleados_ids) > MAX_EMPLEADOS_ZIP:
                return Response({'error': f'Máximo {MAX_EMPLEADOS_ZIP} trabajadores por descarga. Divide la selección en grupos.'}, status=400)

            MAX_LIQUIDACIONES_ZIP = 12
            cantidad_liquidaciones = min(cantidad_liquidaciones, MAX_LIQUIDACIONES_ZIP)

            if not _plan_permite(request.user, 3):
                return Response({'error': 'La descarga masiva de expedientes en ZIP está disponible desde el plan Pyme. Mejora tu suscripción para acceder.'}, status=403)

            empresa = Empresa.objects.get(id=empresa_id, owner=request.user)
            empleados = Empleado.objects.filter(id__in=empleados_ids, empresa=empresa)
            if not empleados.exists():
                return Response({'error': 'No se encontraron trabajadores válidos'}, status=404)

            es_semilla = False  # ya verificado arriba
            meses_corto = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
            zip_buffer = io.BytesIO()

            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                for emp in empleados:
                    rut_limpio = emp.rut.replace("-", "").replace(".", "")
                    carpeta = f"{rut_limpio}_{emp.nombres}_{emp.apellido_paterno}".replace(" ", "_")

                    if 'contrato' in documentos:
                        try:
                            pdf = self._obtener_o_generar_documento(emp, 'contrato', request.user)
                            zip_file.writestr(f"{carpeta}/Contrato.pdf", pdf)
                        except Exception:
                            pass

                    if 'anexo_40h' in documentos:
                        try:
                            pdf = self._obtener_o_generar_documento(emp, 'anexo_40h', request.user)
                            zip_file.writestr(f"{carpeta}/Anexo_Ley_40h.pdf", pdf)
                        except Exception:
                            pass

                    if 'liquidaciones' in documentos and cantidad_liquidaciones > 0:
                        liq_qs = Liquidacion.objects.filter(empleado=emp).order_by('-anio', '-mes')[:cantidad_liquidaciones]
                        for liq in liq_qs:
                            try:
                                pdf = self._obtener_o_generar_documento(emp, f'liquidacion_historica_{liq.mes}_{liq.anio}', request.user)
                                zip_file.writestr(f"{carpeta}/Liquidaciones/Liq_{meses_corto[liq.mes - 1]}_{liq.anio}.pdf", pdf)
                            except Exception:
                                pass

                    if 'amonestaciones' in documentos:
                        for doc in DocumentoLegal.objects.filter(empleado=emp, tipo='AMONESTACION').order_by('fecha_emision'):
                            try:
                                pdf = self._pdf_para_documento_legal(doc, es_semilla)
                                zip_file.writestr(f"{carpeta}/Amonestaciones/Amonestacion_{doc.fecha_emision}.pdf", pdf)
                            except Exception:
                                pass

                    if 'despidos' in documentos:
                        for doc in DocumentoLegal.objects.filter(empleado=emp, tipo='DESPIDO').order_by('fecha_emision'):
                            try:
                                pdf = self._pdf_para_documento_legal(doc, es_semilla)
                                zip_file.writestr(f"{carpeta}/Terminos_Contrato/Termino_{doc.fecha_emision}.pdf", pdf)
                            except Exception:
                                pass

                    if 'mutuo_acuerdo' in documentos:
                        for doc in DocumentoLegal.objects.filter(empleado=emp, tipo='MUTUO_ACUERDO').order_by('fecha_emision'):
                            try:
                                pdf = self._pdf_para_documento_legal(doc, es_semilla)
                                zip_file.writestr(f"{carpeta}/Renuncias/Renuncia_{doc.fecha_emision}.pdf", pdf)
                            except Exception:
                                pass

                    if 'constancias' in documentos:
                        for doc in DocumentoLegal.objects.filter(empleado=emp, tipo='CONSTANCIA').order_by('fecha_emision'):
                            try:
                                pdf = self._pdf_para_documento_legal(doc, es_semilla)
                                zip_file.writestr(f"{carpeta}/Constancias/Constancia_{doc.fecha_emision}.pdf", pdf)
                            except Exception:
                                pass

                    if 'anexos_contrato' in documentos:
                        contrato_emp = Contrato.objects.filter(empleado=emp).first()
                        if contrato_emp:
                            for anexo in AnexoContrato.objects.filter(contrato=contrato_emp).order_by('fecha_emision'):
                                try:
                                    pdf = self._pdf_para_anexo_contrato(anexo, es_semilla)
                                    titulo_corto = anexo.titulo[:30].replace(" ", "_")
                                    zip_file.writestr(f"{carpeta}/Anexos_Contrato/Anexo_{anexo.fecha_emision}_{titulo_corto}.pdf", pdf)
                                except Exception:
                                    pass

            zip_buffer.seek(0)
            nombre_zip = f"Expedientes_{empresa.nombre_legal.replace(' ', '_')}_{datetime.date.today()}.zip"
            response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{nombre_zip}"'
            response['Access-Control-Expose-Headers'] = 'Content-Disposition'
            return response

        except Exception:
            logger.exception('Error al generar ZIP de expedientes')
            return Response({'error': 'Error al generar el ZIP. Inténtalo de nuevo.'}, status=500)
        

    @action(detail=False, methods=['post'])
    def descargar_anexos_zip(self, request):

        if not _plan_permite(request.user, 3):
            return Response(
                {'error': 'La descarga masiva de expedientes en ZIP está disponible desde el plan Pyme. Mejora tu suscripción para acceder.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        empleado_ids = request.data.get('empleados', [])

        if not empleado_ids:
            return Response({'error': 'No se seleccionaron trabajadores'}, status=status.HTTP_400_BAD_REQUEST)

        MAX_EMPLEADOS_ZIP = 50
        if len(empleado_ids) > MAX_EMPLEADOS_ZIP:
            return Response({'error': f'Máximo {MAX_EMPLEADOS_ZIP} trabajadores por descarga. Divide la selección en grupos.'}, status=status.HTTP_400_BAD_REQUEST)
        
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for emp_id in empleado_ids:
                try:
                    empleado = Empleado.objects.get(id=emp_id, empresa__owner=request.user)
                    empresa = empleado.empresa
                    
                    contrato = Contrato.objects.filter(empleado=empleado).first()
                    if not contrato:
                        try: s_base = int(str(empleado.sueldo_base).strip()) if empleado.sueldo_base else 0
                        except: s_base = 0
                        
                        f_inicio = empleado.fecha_ingreso if isinstance(empleado.fecha_ingreso, datetime.date) else datetime.date.today()
                        c_cargo = str(empleado.cargo).strip().upper() if empleado.cargo else 'NO ESPECIFICADO'
                        
                        try:
                            contrato = Contrato.objects.create(
                                empleado=empleado,
                                tipo_contrato='INDEFINIDO',
                                fecha_inicio=f_inicio,
                                sueldo_base=s_base,
                                cargo=c_cargo
                            )
                        except Exception as e:
                            print(f"Aviso BD Contrato - Creando virtual para {empleado.rut}: {e}")
                            class ContratoVirtual:
                                pass
                            contrato = ContratoVirtual()
                            contrato.sueldo_base = s_base

                    meses_zip = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
                    hoy_zip = datetime.date.today()
                    fecha_zip = f"{hoy_zip.day:02d} de {meses_zip[hoy_zip.month - 1]} de {hoy_zip.year}"
                    ciudad_zip = str(
                        getattr(empresa, 'comuna', '') or getattr(empresa, 'ciudad', '') or
                        getattr(empleado, 'comuna', '') or 'Santiago'
                    ).strip().title()
                    context = {
                        'contrato': contrato,
                        'empleado': empleado,
                        'empresa': empresa,
                        'fecha_actual': fecha_zip,
                        'ciudad': ciudad_zip,
                        'es_plan_semilla': _es_plan_semilla(request.user),
                    }

                    template = get_template('anexo_40h.html')
                    html = template.render(context)
                    
                    pdf_buffer = io.BytesIO()
                    pisa_status = pisa.CreatePDF(html, dest=pdf_buffer)
                    
                    if not pisa_status.err:
                        nombre_archivo = f"Anexo_40h_{empleado.rut}.pdf"
                        zip_file.writestr(nombre_archivo, pdf_buffer.getvalue())
                        
                except Exception as e:
                    print(f"Error fatal saltando empleado {emp_id} en ZIP: {e}")
                    continue 
        
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = 'attachment; filename="Anexos_Masivos_40h.zip"'
        return response

    @action(detail=True, methods=['post'], url_path='digitalizar_contrato')
    def digitalizar_contrato(self, request, pk=None):
        from .extractor_contrato import extraer_campos_contrato

        archivo = request.FILES.get('file')
        if not archivo:
            return Response({'error': 'No se recibió ningún archivo.'}, status=400)

        MIME_PERMITIDOS = {
            'application/pdf': 'application/pdf',
            'image/jpeg':      'image/jpeg',
            'image/png':       'image/png',
        }
        mime = archivo.content_type or ''
        if mime not in MIME_PERMITIDOS:
            return Response(
                {'error': 'Formato no soportado. Sube un PDF, JPG o PNG.'},
                status=400,
            )

        LIMITE_BYTES = 20 * 1024 * 1024  # 20 MB
        if archivo.size > LIMITE_BYTES:
            return Response({'error': 'El archivo supera el límite de 20 MB.'}, status=400)

        try:
            campos = extraer_campos_contrato(archivo.read(), mime)
        except RuntimeError as e:
            return Response({'error': str(e)}, status=502)
        except Exception as e:
            return Response(
                {'error': f'Error inesperado al analizar el documento: {e}'},
                status=500,
            )

        return Response(campos)


class ContratoViewSet(viewsets.ModelViewSet):
    serializer_class = ContratoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Contrato.objects.filter(empleado__empresa__owner=self.request.user)
        empleado_id = self.request.query_params.get('empleado')
        if empleado_id:
            queryset = queryset.filter(empleado_id=empleado_id)
        return queryset

    def _build_contrato_context(self, contrato, es_plan_semilla):
        return _ctx_contrato(contrato, es_plan_semilla)

    @action(detail=True, methods=['post'])
    def generar_contrato_pdf(self, request, pk=None):
        try:
            contrato = self.get_object()
            es_plan_semilla = _es_plan_semilla(request.user)
            context = self._build_contrato_context(contrato, es_plan_semilla)
            html = get_template('contrato_trabajo.html').render(context)
            pdf_buf = io.BytesIO()
            pisa_status = pisa.CreatePDF(html, dest=pdf_buf)
            if pisa_status.err:
                return Response({'error': 'Error al generar el PDF.'}, status=500)
            if contrato.archivo_contrato:
                contrato.archivo_contrato.delete(save=False)
            nombre = f"Contrato_{contrato.empleado.rut}.pdf"
            contrato.archivo_contrato.save(nombre, ContentFile(pdf_buf.getvalue()), save=True)
            return Response({'ok': True, 'mensaje': 'Contrato generado y guardado exitosamente.'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def descargar_contrato(self, request, pk=None):
        try:
            contrato = self.get_object()
            if not contrato.archivo_contrato:
                return Response({'error': 'El contrato aún no tiene PDF generado.'}, status=status.HTTP_404_NOT_FOUND)
            nombre = f"Contrato_{contrato.empleado.rut}.pdf"
            response = HttpResponse(contrato.archivo_contrato.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{nombre}"'
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'])
    def generar_anexo_40h(self, request, pk=None):
        try:
            contrato = self.get_object()
            es_plan_semilla = _es_plan_semilla(request.user)
            context = self._build_contrato_context(contrato, es_plan_semilla)
            html = get_template('anexo_40h.html').render(context)
            pdf_buf = io.BytesIO()
            pisa_status = pisa.CreatePDF(html, dest=pdf_buf)
            if pisa_status.err:
                return Response({'error': 'Error al generar el PDF.'}, status=500)
            if contrato.archivo_anexo_40h:
                contrato.archivo_anexo_40h.delete(save=False)
            nombre = f"Anexo_40h_{contrato.empleado.rut}.pdf"
            contrato.archivo_anexo_40h.save(nombre, ContentFile(pdf_buf.getvalue()), save=True)
            return Response({'ok': True, 'mensaje': 'Anexo 40h generado y guardado exitosamente.'})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def descargar_anexo_40h(self, request, pk=None):
        try:
            contrato = self.get_object()
            if not contrato.archivo_anexo_40h:
                return Response({'error': 'El anexo 40h aún no tiene PDF generado.'}, status=status.HTTP_404_NOT_FOUND)
            nombre = f"Anexo_40h_{contrato.empleado.rut}.pdf"
            response = HttpResponse(contrato.archivo_anexo_40h.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{nombre}"'
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # Mantener compatibilidad con descarga masiva ZIP (sin guardar)
    @action(detail=True, methods=['get'])
    def generar_anexo(self, request, pk=None):
        try:
            contrato = self.get_object()
            es_plan_semilla = _es_plan_semilla(request.user)
            context = self._build_contrato_context(contrato, es_plan_semilla)
            html = get_template('anexo_40h.html').render(context)
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="Anexo_40h_{contrato.empleado.rut}.pdf"'
            pisa.CreatePDF(html, dest=response)
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
# ==========================================
# LOGIN CON RATE LIMITING
# ==========================================
from dj_rest_auth.views import LoginView as DjRestLoginView

class ThrottledLoginView(DjRestLoginView):
    throttle_classes = [LoginRateThrottle]

from dj_rest_auth.views import PasswordResetView as DjRestPasswordResetView

class ThrottledPasswordResetView(DjRestPasswordResetView):
    throttle_classes = [PasswordResetRateThrottle]

# ==========================================
# REGISTRO DE NUEVOS CLIENTES
# ==========================================
@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([RegisterRateThrottle])
def registrar_cliente(request):
    rut = request.data.get('rut')
    password = request.data.get('password')
    # Atrapamos el correo (por si tu React lo manda como 'email' o como 'correo')
    email = request.data.get('email') or request.data.get('correo')
    first_name=request.data.get('nombres', ''),
    last_name=request.data.get('apellidos', '')
    
    # Validaciones básicas
    if not rut or not password or not email:
        return Response({'error': 'Faltan datos obligatorios (RUT, contraseña o correo)'}, status=400)

    try:
        with transaction.atomic():
            # 1. Creamos el acceso en la tabla core_users (User de Django)
            user = User.objects.create_user(
                username=rut, 
                password=password,
                email=email
            )

            plan_semilla, creado = Plan.objects.get_or_create(
                nombre='Semilla',
                defaults={
                    'max_empresas': 1,
                    'limite_trabajadores': 5,
                    'precio': 0,
                    'nivel': 1,
                    'activo': True,
                }
            )
            if not creado and plan_semilla.nivel != 1:
                plan_semilla.nivel = 1
                plan_semilla.save(update_fields=['nivel'])
            
            # 2. Creamos el perfil en core_cliente 
            cliente = Cliente.objects.create(
                usuario=user,
                rut=rut,
                correo=email,
                plan=plan_semilla  # Asignamos el plan "Semilla" por defecto (usando el objeto obtenido o creado arriba
            )
            
           
        return Response({'mensaje': 'Cliente creado con éxito'}, status=201)
        
    except IntegrityError:
        return Response({'error': 'Este RUT ya está registrado en el sistema.'}, status=400)
    except Exception as e:
        return Response({'error': str(e)}, status=500)
    
class AnexoContratoViewSet(viewsets.ModelViewSet):
    serializer_class = AnexoContratoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = AnexoContrato.objects.filter(
            contrato__empleado__empresa__owner=self.request.user
        ).order_by('-fecha_emision')
        empleado_id = self.request.query_params.get('empleado')
        if empleado_id:
            queryset = queryset.filter(contrato__empleado_id=empleado_id)
        return queryset

    def create(self, request, *args, **kwargs):
        contrato_id = request.data.get('contrato')
        try:
            contrato = Contrato.objects.get(id=contrato_id, empleado__empresa__owner=request.user)
        except Contrato.DoesNotExist:
            return Response({'error': 'Contrato no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
        return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        anexo = serializer.save()
        try:
            contrato = anexo.contrato
            empleado = contrato.empleado
            empresa = empleado.empresa
            es_plan_semilla = _es_plan_semilla(self.request.user)
            meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
            fecha_obj = anexo.fecha_emision
            fecha_espanol = f"{fecha_obj.day:02d} de {meses[fecha_obj.month - 1]} de {fecha_obj.year}"
            ciudad = str(getattr(empresa, 'comuna', '') or getattr(empresa, 'ciudad', '') or 'Santiago').strip().title()
            context = {
                'anexo': anexo, 'contrato': contrato, 'empleado': empleado,
                'empresa': empresa, 'fecha_actual': fecha_espanol,
                'ciudad': ciudad, 'es_plan_semilla': es_plan_semilla,
            }
            html = get_template('anexo_contrato.html').render(context)
            pdf_buf = io.BytesIO()
            pisa_status = pisa.CreatePDF(html, dest=pdf_buf)
            if not pisa_status.err and pdf_buf.getvalue():
                nombre = f"AnexoContrato_{empleado.rut}_{anexo.fecha_emision}.pdf"
                anexo.archivo_pdf.save(nombre, ContentFile(pdf_buf.getvalue()), save=True)
        except Exception:
            pass  # No bloqueamos el guardado si el PDF falla

    @action(detail=True, methods=['get'])
    def generar_pdf(self, request, pk=None):
        try:
            anexo = self.get_object()
            contrato = anexo.contrato
            empleado = contrato.empleado
            empresa = empleado.empresa
            es_plan_semilla = _es_plan_semilla(request.user)

            meses = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]
            hoy = anexo.fecha_emision
            fecha_espanol = f"{hoy.day:02d} de {meses[hoy.month - 1]} de {hoy.year}"
            ciudad = str(getattr(empresa, 'comuna', '') or getattr(empresa, 'ciudad', '') or 'Santiago').strip().title()

            context = {
                'anexo': anexo,
                'contrato': contrato,
                'empleado': empleado,
                'empresa': empresa,
                'fecha_actual': fecha_espanol,
                'ciudad': ciudad,
                'es_plan_semilla': es_plan_semilla,
            }
            template = get_template('anexo_contrato.html')
            html = template.render(context)

            response = HttpResponse(content_type='application/pdf')
            nombre = f"Anexo_{empleado.rut}_{hoy}.pdf"
            response['Content-Disposition'] = f'attachment; filename="{nombre}"'
            pisa_status = pisa.CreatePDF(html, dest=response)
            if pisa_status.err:
                return Response({'error': 'Error generando PDF'}, status=500)
            return response
        except Exception as e:
            return Response({'error': str(e)}, status=500)


class LiquidacionViewSet(viewsets.ModelViewSet):
    queryset = Liquidacion.objects.all().order_by('-anio', '-mes')
    serializer_class = LiquidacionSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        # Solo liquidaciones de empleados que pertenecen al usuario autenticado
        queryset = Liquidacion.objects.filter(
            empleado__empresa__owner=self.request.user
        ).order_by('-anio', '-mes')
        empleado_id = self.request.query_params.get('empleado', None)
        if empleado_id is not None:
            queryset = queryset.filter(empleado_id=empleado_id)
        return queryset

    def create(self, request, *args, **kwargs):
        data = request.data
        empleado_id = data.get('empleado')

        try:
            # Validar que el empleado pertenezca al usuario autenticado
            empleado = Empleado.objects.get(id=empleado_id, empresa__owner=request.user)
            contrato = Contrato.objects.filter(empleado=empleado).first()
            
            if not contrato:
                return Response({'error': 'El trabajador no tiene un contrato activo.'}, status=status.HTTP_400_BAD_REQUEST)

            # 1. ASISTENCIA
            dias_trabajados = int(data.get('dias_trabajados', 30))
            dias_ausencia = int(data.get('dias_ausencia', 0))
            dias_licencia = int(data.get('dias_licencia', 0))
            dias_no_contratados = int(data.get('dias_no_contratados', 0))
            
            # Los días a pagar de sueldo base son 30 menos las ausencias y licencias
            dias_a_pagar = 30 - dias_ausencia - dias_licencia - dias_no_contratados
            if dias_a_pagar < 0: dias_a_pagar = 0

            # 2. ARREGLOS DINÁMICOS (JSON)
            detalle_imponibles = data.get('detalle_haberes_imponibles', [])
            detalle_no_imponibles = data.get('detalle_haberes_no_imponibles', [])
            detalle_horas_extras = data.get('detalle_horas_extras', [])
            detalle_otros_descuentos = data.get('detalle_otros_descuentos', [])

            # Sumas matemáticas de los arreglos
            suma_imponibles_extra = sum(int(item.get('valor', 0)) for item in detalle_imponibles)
            suma_horas_extras = sum(int(item.get('valor', 0)) for item in detalle_horas_extras)
            suma_no_imponibles = sum(int(item.get('valor', 0)) for item in detalle_no_imponibles)
            suma_otros_descuentos = sum(int(item.get('valor', 0)) for item in detalle_otros_descuentos)

            # 3. CÁLCULO DE HABERES
            sueldo_base_mensual = contrato.sueldo_base
            sueldo_base_proporcional = math.floor((sueldo_base_mensual / 30) * dias_a_pagar)
            
            # Gratificación
            tope_gratificacion = 200000 
            base_gratificacion = sueldo_base_proporcional + suma_imponibles_extra + suma_horas_extras
            gratificacion_calculada = math.floor(base_gratificacion * 0.25)
            gratificacion_final = min(gratificacion_calculada, tope_gratificacion) if contrato.gratificacion_legal == 'MENSUAL' else 0

            total_imponible = base_gratificacion + gratificacion_final
            total_haberes = total_imponible + suma_no_imponibles

            # 4. CÁLCULO DE DESCUENTOS LEGALES
            tasas_afp = {
                'MODELO': 0.1058, 'HABITAT': 0.1127, 'PROVIDA': 0.1145,
                'CAPITAL': 0.1144, 'CUPRUM': 0.1144, 'PLANVITAL': 0.1116, 'UNO': 0.1049
            }
            nombre_afp = (empleado.afp or 'MODELO').upper()
            tasa_afp = tasas_afp.get(nombre_afp, 0.11)
            afp_monto = math.floor(total_imponible * tasa_afp)

            # Salud (Isapre UF vs Fonasa 7%)
            salud_nombre = (empleado.sistema_salud or 'FONASA').upper()
            if salud_nombre == 'ISAPRE' and empleado.plan_isapre_uf > 0:
                # Simulación valor UF 
                VALOR_UF = 38000 
                salud_monto = math.floor(float(empleado.plan_isapre_uf) * VALOR_UF)
                isapre_uf = empleado.plan_isapre_uf
                # La ley exige descontar al menos el 7%, si el plan UF es menor, se cobra 7%
                minimo_legal = math.floor(total_imponible * 0.07)
                if salud_monto < minimo_legal:
                    salud_monto = minimo_legal
            else:
                salud_monto = math.floor(total_imponible * 0.07)
                isapre_uf = 0

            # Seguro Cesantía
            seguro_cesantia = math.floor(total_imponible * 0.006) if contrato.tipo_contrato == 'INDEFINIDO' else 0

            # Quincena y otros
            anticipo_quincena = contrato.monto_quincena if contrato.tiene_quincena else 0
            total_descuentos = afp_monto + salud_monto + seguro_cesantia + anticipo_quincena + suma_otros_descuentos

            # 5. SUELDO LÍQUIDO FINAL
            sueldo_liquido = total_haberes - total_descuentos

            # 6. GUARDAR EN BASE DE DATOS
            liquidacion = Liquidacion.objects.create(
                empleado=empleado, mes=data.get('mes'), anio=data.get('anio'),
                dias_trabajados=dias_trabajados, dias_licencia=dias_licencia,
                dias_ausencia=dias_ausencia, dias_no_contratados=dias_no_contratados,
                sueldo_base=sueldo_base_proporcional, gratificacion=gratificacion_final,
                detalle_haberes_imponibles=detalle_imponibles,
                detalle_horas_extras=detalle_horas_extras,
                detalle_haberes_no_imponibles=detalle_no_imponibles,
                detalle_otros_descuentos=detalle_otros_descuentos,
                afp_nombre=nombre_afp, afp_monto=afp_monto,
                salud_nombre=salud_nombre, isapre_cotizacion_uf=isapre_uf, salud_monto=salud_monto,
                seguro_cesantia=seguro_cesantia, anticipo_quincena=anticipo_quincena,
                total_imponible=total_imponible, total_haberes=total_haberes,
                total_descuentos=total_descuentos, sueldo_liquido=sueldo_liquido
            )

            serializer = self.get_serializer(liquidacion)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        except Empleado.DoesNotExist:
            return Response({'error': 'Trabajador no encontrado o no autorizado.'}, status=status.HTTP_404_NOT_FOUND)
        except IntegrityError:
            return Response(
                {'error': f'Ya existe una liquidación para este trabajador en el período indicado.'},
                status=status.HTTP_409_CONFLICT
            )
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'])
    def generar_pdf(self, request, pk=None):
        try:
            liquidacion = self.get_object()
            es_plan_semilla = _es_plan_semilla(request.user)
            empleado = liquidacion.empleado
            empresa = empleado.empresa
            contrato = Contrato.objects.filter(empleado=empleado).first()

            meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio", "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
            mes_nombre = meses[liquidacion.mes - 1]

            # Transformar número a palabras (Ej: 542000 -> "quinientos cuarenta y dos mil")
            sueldo_seguro = int(liquidacion.sueldo_liquido or 0)
            liquido_palabras = num2words(sueldo_seguro, lang='es')

            det_no_imp = liquidacion.detalle_haberes_no_imponibles
            if not isinstance(det_no_imp, list): det_no_imp = []
            suma_no_imponibles = sum(int(item.get('valor', 0)) for item in det_no_imp if isinstance(item, dict))

            det_otros_dsctos = liquidacion.detalle_otros_descuentos
            if not isinstance(det_otros_dsctos, list): det_otros_dsctos = []
            suma_otros_descuentos = sum(int(item.get('valor', 0)) for item in det_otros_dsctos if isinstance(item, dict))
            
            total_ley = (liquidacion.afp_monto or 0) + (liquidacion.salud_monto or 0) + (liquidacion.seguro_cesantia or 0) + (liquidacion.impuesto_unico or 0)
            total_otros_dsctos = (liquidacion.anticipo_quincena or 0) + suma_otros_descuentos

            context = {
                'liquidacion': liquidacion,
                'empleado': empleado,
                'empresa': empresa,
                'contrato': contrato,
                'mes_nombre': mes_nombre.upper(),
                'liquido_palabras': liquido_palabras,
                'total_no_imponible': suma_no_imponibles,
                'total_ley': total_ley,
                'total_otros_dsctos': total_otros_dsctos,
                'es_plan_semilla': es_plan_semilla
            }

            template = get_template('liquidacion.html')
            html = template.render(context)

            response = HttpResponse(content_type='application/pdf')
            nombre_archivo = f'Liquidacion_{liquidacion.mes}_{liquidacion.anio}_{empleado.rut}.pdf'
            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            pisa_status = pisa.CreatePDF(html, dest=response)

            if pisa_status.err:
                print("Error interno de pisa (xhtml2pdf)")
                return Response({'error': 'Error al generar PDF'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
            
            return response

        except Exception as e:
            logger.exception('Error al generar PDF de liquidación')
            return Response({'error': f'Error generando PDF: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='exportar_previred')
    def exportar_previred(self, request):
        if not _plan_permite(request.user, 3):
            return Response(
                {'error': 'La exportación Previred está disponible desde el plan Pyme. Mejora tu suscripción para acceder.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        mes_param = request.query_params.get('mes')
        anio_param = request.query_params.get('anio')
        empresa_id = request.query_params.get('empresa')

        if not mes_param or not anio_param:
            return Response({'error': 'Se requieren los parámetros mes y anio.'}, status=400)
        try:
            mes = int(mes_param)
            anio = int(anio_param)
        except ValueError:
            return Response({'error': 'Parámetros mes y anio deben ser numéricos.'}, status=400)

        qs = Liquidacion.objects.filter(
            empleado__empresa__owner=request.user,
            mes=mes, anio=anio,
        ).select_related('empleado', 'empleado__empresa')

        if empresa_id:
            qs = qs.filter(empleado__empresa_id=empresa_id)

        if not qs.exists():
            return Response({'error': 'No hay liquidaciones para el período seleccionado.'}, status=404)

        lineas = []
        for liq in qs:
            emp = liq.empleado
            empresa = emp.empresa
            contrato = Contrato.objects.filter(empleado=emp).first()

            # ── Identificación trabajador ──────────────────────────────────
            rut_num, rut_dv = _rut_partes(emp.rut)
            apellido_m = emp.apellido_materno or ''
            sexo_cod = '2' if emp.sexo == 'F' else '1'
            fecha_nac = _fmt_fecha_previred(emp.fecha_nacimiento)
            fecha_ing = _fmt_fecha_previred(emp.fecha_ingreso)
            tipo_trab = '01'
            nac_cod = '152'  # Chile

            # ── Contrato ───────────────────────────────────────────────────
            dias_trab = str(int(liq.dias_trabajados or 30))
            tipo_ctto = _TIPO_CONTRATO_PREVIRED.get(
                contrato.tipo_contrato if contrato else 'INDEFINIDO', '1'
            )
            es_indefinido = contrato.tipo_contrato == 'INDEFINIDO' if contrato else False
            movimiento = '0'  # vigente
            rut_emp_num, rut_emp_dv = _rut_partes(empresa.rut)

            # ── AFP ────────────────────────────────────────────────────────
            nombre_afp = (liq.afp_nombre or 'MODELO').upper()
            cod_afp = _AFP_CODIGOS_PREVIRED.get(nombre_afp, '08')
            renta_imp = int(liq.total_imponible or 0)
            cotiz_afp = str(int(liq.afp_monto or 0))
            sis = str(math.floor(renta_imp * _TASA_SIS))

            # ── Salud ──────────────────────────────────────────────────────
            sistema = (liq.salud_nombre or 'FONASA').upper()
            if sistema == 'FONASA':
                cod_salud = '00'
            else:
                cod_salud = _ISAPRE_CODIGOS_PREVIRED.get(sistema, '00')
            cotiz_salud = str(int(liq.salud_monto or 0))
            uf_isapre = str(float(liq.isapre_cotizacion_uf or 0))

            # ── Mutual AT/EP ───────────────────────────────────────────────
            cotiz_mutual = str(math.floor(renta_imp * _TASA_MUTUAL_AT))

            # ── AFC Cesantía ───────────────────────────────────────────────
            ind_afc = '1' if es_indefinido else '0'
            cotiz_afc_trab = str(int(liq.seguro_cesantia or 0))
            if es_indefinido:
                cotiz_afc_emp = str(math.floor(renta_imp * _TASA_AFC_EMP_INDEFINIDO))
            elif contrato and contrato.tipo_contrato == 'PLAZO_FIJO':
                cotiz_afc_emp = str(math.floor(renta_imp * _TASA_AFC_EMP_PLAZO))
            else:
                cotiz_afc_emp = '0'
            renta_imp_afc = str(renta_imp) if es_indefinido else '0'

            # ── Reforma 2025 ───────────────────────────────────────────────
            tipo_jornada_code = _TIPO_JORNADA_PREVIRED.get(
                contrato.tipo_jornada if contrato else 'ORDINARIA', '1'
            )
            cotiz_expectativa = str(math.floor(renta_imp * _TASA_EXPECTATIVA_VIDA))

            # ── Construir array de 105 campos (base cero) ──────────────────
            campos = ['0'] * 105

            # Trabajador / contrato (campos 1-17, índices 0-16)
            campos[0]  = rut_num
            campos[1]  = rut_dv
            campos[2]  = emp.apellido_paterno
            campos[3]  = apellido_m
            campos[4]  = emp.nombres
            campos[5]  = sexo_cod
            campos[6]  = fecha_nac
            campos[7]  = nac_cod
            campos[8]  = tipo_trab
            campos[9]  = fecha_ing
            campos[10] = ''   # fecha término (activo)
            campos[11] = ''   # causal término
            campos[12] = dias_trab
            campos[13] = tipo_ctto
            campos[14] = movimiento
            campos[15] = rut_emp_num
            campos[16] = rut_emp_dv
            # índices 17-23: padding → '0' (ya inicializados)

            # AFP (campos 25-28, índices 24-27)
            campos[24] = cod_afp
            campos[25] = str(renta_imp)
            campos[26] = cotiz_afp
            campos[27] = sis
            # índices 28-43: extras AFP → '0'

            # Salud (campos 45-48, índices 44-47)
            campos[44] = cod_salud
            campos[45] = str(renta_imp)
            campos[46] = cotiz_salud
            campos[47] = uf_isapre
            # índices 48-59: extras salud → '0'

            # Mutual AT/EP (campos 61-63, índices 60-62)
            campos[60] = _MUTUAL_DEFAULT
            campos[61] = str(renta_imp)
            campos[62] = cotiz_mutual
            # índices 63-69: extras mutual → '0'

            # AFC (campos 71-74, índices 70-73)
            campos[70] = ind_afc
            campos[71] = renta_imp_afc
            campos[72] = cotiz_afc_trab
            campos[73] = cotiz_afc_emp
            # índices 74-84: extras AFC → '0'

            # Reforma 2025 (campos 86-88, índices 85-87)
            campos[85] = '0'               # RIMA
            campos[86] = tipo_jornada_code # tipo jornada ley 40h
            campos[87] = cotiz_expectativa # expectativa de vida 0.9%
            # índices 88-104: extras → '0'

            lineas.append(';'.join(campos))

        meses_nombres = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
        ]
        nombre_archivo = f'Previred_{meses_nombres[mes - 1]}_{anio}.txt'
        contenido = '\n'.join(lineas)
        response = HttpResponse(contenido, content_type='text/plain; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'
        return response

    @action(detail=False, methods=['get'], url_path='libro_remuneraciones')
    def libro_remuneraciones(self, request):
        if not _plan_permite(request.user, 3):
            return Response(
                {'error': 'El Libro de Remuneraciones está disponible desde el plan Pyme. Mejora tu suscripción para acceder.'},
                status=status.HTTP_403_FORBIDDEN,
            )
        mes_param  = request.query_params.get('mes')
        anio_param = request.query_params.get('anio')
        empresa_id = request.query_params.get('empresa')
        formato    = request.query_params.get('formato', 'excel')

        if not mes_param or not anio_param:
            return Response({'error': 'Se requieren los parámetros mes y anio.'}, status=400)
        try:
            mes  = int(mes_param)
            anio = int(anio_param)
        except ValueError:
            return Response({'error': 'Parámetros mes y anio deben ser numéricos.'}, status=400)

        qs = Liquidacion.objects.filter(
            empleado__empresa__owner=request.user,
            mes=mes, anio=anio,
        ).select_related('empleado', 'empleado__empresa').order_by(
            'empleado__ficha_numero', 'empleado__apellido_paterno'
        )
        if empresa_id:
            qs = qs.filter(empleado__empresa_id=empresa_id)
        if not qs.exists():
            return Response({'error': 'No hay liquidaciones para el período seleccionado.'}, status=404)

        empresa = qs.first().empleado.empresa
        meses_nombres = [
            'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
            'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre',
        ]
        mes_nombre = meses_nombres[mes - 1]

        # ── Helper formato CLP para el PDF ────────────────────────────────────
        def clp(n):
            if not n:
                return '$0'
            return f'${int(n):,}'.replace(',', '.')

        # ── Preparación de filas (compartido por Excel y PDF) ─────────────────
        filas = []
        totales = {k: 0 for k in [
            'sueldo_base', 'gratificacion', 'otros_imp', 'total_imponible',
            'no_imponibles', 'total_haberes', 'cotiz_afp', 'cotiz_salud',
            'cesantia', 'imp_unico', 'anticipo', 'otros_desc',
            'total_descuentos', 'sueldo_liquido',
        ]}

        for liq in qs:
            emp = liq.empleado
            det_imp = liq.detalle_haberes_imponibles or []
            if not isinstance(det_imp, list): det_imp = []
            det_hex = liq.detalle_horas_extras or []
            if not isinstance(det_hex, list): det_hex = []
            det_noi = liq.detalle_haberes_no_imponibles or []
            if not isinstance(det_noi, list): det_noi = []
            det_odc = liq.detalle_otros_descuentos or []
            if not isinstance(det_odc, list): det_odc = []

            otros_imp  = sum(int(d.get('valor', 0)) for d in det_imp if isinstance(d, dict))
            otros_imp += sum(int(d.get('valor', 0)) for d in det_hex if isinstance(d, dict))
            no_impon   = sum(int(d.get('valor', 0)) for d in det_noi if isinstance(d, dict))
            otros_desc = sum(int(d.get('valor', 0)) for d in det_odc if isinstance(d, dict))

            filas.append({
                'ficha':        emp.ficha_numero or '',
                'rut':          emp.rut,
                'nombre':       f'{emp.apellido_paterno} {emp.apellido_materno or ""} {emp.nombres}'.strip(),
                'cargo':        emp.cargo or '',
                'dias':         liq.dias_trabajados,
                'sueldo_base':  int(liq.sueldo_base),
                'gratificacion':int(liq.gratificacion),
                'otros_imp':    otros_imp,
                'total_imp':    int(liq.total_imponible),
                'no_imp':       no_impon,
                'total_hab':    int(liq.total_haberes),
                'afp_nombre':   liq.afp_nombre or '',
                'cotiz_afp':    int(liq.afp_monto),
                'salud_nombre': liq.salud_nombre or '',
                'cotiz_salud':  int(liq.salud_monto),
                'cesantia':     int(liq.seguro_cesantia),
                'imp_unico':    int(liq.impuesto_unico or 0),
                'anticipo':     int(liq.anticipo_quincena or 0),
                'otros_desc':   otros_desc,
                'total_desc':   int(liq.total_descuentos),
                'sueldo_liq':   int(liq.sueldo_liquido),
            })

            totales['sueldo_base']      += int(liq.sueldo_base)
            totales['gratificacion']    += int(liq.gratificacion)
            totales['otros_imp']        += otros_imp
            totales['total_imponible']  += int(liq.total_imponible)
            totales['no_imponibles']    += no_impon
            totales['total_haberes']    += int(liq.total_haberes)
            totales['cotiz_afp']        += int(liq.afp_monto)
            totales['cotiz_salud']      += int(liq.salud_monto)
            totales['cesantia']         += int(liq.seguro_cesantia)
            totales['imp_unico']        += int(liq.impuesto_unico or 0)
            totales['anticipo']         += int(liq.anticipo_quincena or 0)
            totales['otros_desc']       += otros_desc
            totales['total_descuentos'] += int(liq.total_descuentos)
            totales['sueldo_liquido']   += int(liq.sueldo_liquido)

        nombre_base = f'LibroRemuneraciones_{mes_nombre}_{anio}_{empresa.rut}'

        # ══════════════════════════════════════════════════════════════════════
        # RAMA PDF
        # ══════════════════════════════════════════════════════════════════════
        if formato == 'pdf':
            # Pre-formatear montos para el template
            for f in filas:
                for k in ['sueldo_base','gratificacion','otros_imp','total_imp','no_imp',
                          'total_hab','cotiz_afp','cotiz_salud','cesantia','imp_unico',
                          'anticipo','otros_desc','total_desc','sueldo_liq']:
                    f[f'{k}_fmt'] = clp(f[k])

            totales_fmt = {k: clp(v) for k, v in totales.items()}

            context = {
                'empresa':      empresa,
                'mes_nombre':   mes_nombre,
                'anio':         anio,
                'filas':        filas,
                'totales':      totales,
                'totales_fmt':  totales_fmt,
                'fecha_emision': timezone.now().date().strftime('%d/%m/%Y'),
                'n_trabajadores': len(filas),
            }
            template = get_template('libro_remuneraciones.html')
            html = template.render(context)

            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{nombre_base}.pdf"'
            pisa_status = pisa.CreatePDF(html, dest=response)
            if pisa_status.err:
                return Response({'error': 'Error al generar PDF'}, status=500)
            return response

        # ══════════════════════════════════════════════════════════════════════
        # RAMA EXCEL (sin cambios respecto al paso 1)
        # ══════════════════════════════════════════════════════════════════════

        # ── Estilos ───────────────────────────────────────────────────────────
        COLOR_HEADER   = '1E3A5F'
        COLOR_TOTALES  = 'F59E0B'
        COLOR_FILA_PAR = 'F1F5F9'

        ft_titulo    = Font(name='Calibri', bold=True, size=14, color='1E3A5F')
        ft_subtit    = Font(name='Calibri', bold=True, size=11, color='334155')
        ft_header    = Font(name='Calibri', bold=True, size=9,  color='FFFFFF')
        ft_dato      = Font(name='Calibri', size=9)
        ft_total     = Font(name='Calibri', bold=True, size=9)
        ft_total_liq = Font(name='Calibri', bold=True, size=9, color='7C3AED')

        al_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        al_left   = Alignment(horizontal='left',   vertical='center')
        al_right  = Alignment(horizontal='right',  vertical='center')

        fill_header = PatternFill('solid', fgColor=COLOR_HEADER)
        fill_total  = PatternFill('solid', fgColor=COLOR_TOTALES)
        fill_par    = PatternFill('solid', fgColor=COLOR_FILA_PAR)

        thin  = Side(style='thin', color='CBD5E1')
        borde = Border(left=thin, right=thin, top=thin, bottom=thin)

        FMT_CLP  = '#,##0'
        FMT_TEXT = '@'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Libro {mes_nombre} {anio}'
        ws.sheet_view.showGridLines = False

        COLS = [
            ('N°',                    5),
            ('RUT',                  12),
            ('Apellidos y Nombres',  28),
            ('Cargo',                16),
            ('Días\nTrab.',           7),
            ('Sueldo\nBase',         12),
            ('Gratif.',              11),
            ('Otros Hab.\nImpon.',   12),
            ('Total\nImponible',     13),
            ('Hab. No\nImponibles',  13),
            ('Total\nHaberes',       12),
            ('AFP',                   9),
            ('Cotiz.\nAFP',          11),
            ('Salud',                 9),
            ('Cotiz.\nSalud',        11),
            ('Cesantía',             10),
            ('Imp.\nÚnico',          10),
            ('Anticipo',             10),
            ('Otros\nDesc.',         10),
            ('Total\nDescuentos',    13),
            ('Alcance\nLíquido',     13),
        ]
        for i, (_, ancho) in enumerate(COLS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = ancho

        N_COLS     = len(COLS)
        ultima_col = get_column_letter(N_COLS)

        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 16

        ws.merge_cells(f'A1:{ultima_col}1')
        c = ws['A1']
        c.value     = empresa.nombre_legal.upper()
        c.font      = ft_titulo
        c.alignment = al_center

        ws.merge_cells(f'A2:{ultima_col}2')
        c = ws['A2']
        c.value     = 'LIBRO DE REMUNERACIONES'
        c.font      = ft_subtit
        c.alignment = al_center

        ws.merge_cells(f'A3:{ultima_col}3')
        c = ws['A3']
        c.value     = f'RUT: {empresa.rut}     Período: {mes_nombre} {anio}'
        c.font      = Font(name='Calibri', size=10, color='475569')
        c.alignment = al_center

        ws.row_dimensions[4].height = 6

        ws.row_dimensions[5].height = 36
        for col_idx, (label, _) in enumerate(COLS, start=1):
            c = ws.cell(row=5, column=col_idx, value=label)
            c.font      = ft_header
            c.fill      = fill_header
            c.alignment = al_center
            c.border    = borde

        for fila_idx, f in enumerate(filas, start=6):
            fill_fila = fill_par if fila_idx % 2 == 0 else None
            valores = [
                (f['ficha'],         FMT_TEXT, al_center),
                (f['rut'],           FMT_TEXT, al_left),
                (f['nombre'],        FMT_TEXT, al_left),
                (f['cargo'],         FMT_TEXT, al_left),
                (f['dias'],          FMT_TEXT, al_center),
                (f['sueldo_base'],   FMT_CLP,  al_right),
                (f['gratificacion'], FMT_CLP,  al_right),
                (f['otros_imp'],     FMT_CLP,  al_right),
                (f['total_imp'],     FMT_CLP,  al_right),
                (f['no_imp'],        FMT_CLP,  al_right),
                (f['total_hab'],     FMT_CLP,  al_right),
                (f['afp_nombre'],    FMT_TEXT, al_center),
                (f['cotiz_afp'],     FMT_CLP,  al_right),
                (f['salud_nombre'],  FMT_TEXT, al_center),
                (f['cotiz_salud'],   FMT_CLP,  al_right),
                (f['cesantia'],      FMT_CLP,  al_right),
                (f['imp_unico'],     FMT_CLP,  al_right),
                (f['anticipo'],      FMT_CLP,  al_right),
                (f['otros_desc'],    FMT_CLP,  al_right),
                (f['total_desc'],    FMT_CLP,  al_right),
                (f['sueldo_liq'],    FMT_CLP,  al_right),
            ]
            ws.row_dimensions[fila_idx].height = 15
            for col_idx, (valor, fmt, alin) in enumerate(valores, start=1):
                c = ws.cell(row=fila_idx, column=col_idx, value=valor)
                c.number_format = fmt
                c.font          = ft_dato
                c.alignment     = alin
                c.border        = borde
                if fill_fila:
                    c.fill = fill_fila

        fila_total = len(filas) + 6
        ws.row_dimensions[fila_total].height = 18
        vals_total = [
            ('', FMT_TEXT), ('', FMT_TEXT),
            ('TOTALES', FMT_TEXT), ('', FMT_TEXT), ('', FMT_TEXT),
            (totales['sueldo_base'],      FMT_CLP),
            (totales['gratificacion'],    FMT_CLP),
            (totales['otros_imp'],        FMT_CLP),
            (totales['total_imponible'],  FMT_CLP),
            (totales['no_imponibles'],    FMT_CLP),
            (totales['total_haberes'],    FMT_CLP),
            ('', FMT_TEXT),
            (totales['cotiz_afp'],        FMT_CLP),
            ('', FMT_TEXT),
            (totales['cotiz_salud'],      FMT_CLP),
            (totales['cesantia'],         FMT_CLP),
            (totales['imp_unico'],        FMT_CLP),
            (totales['anticipo'],         FMT_CLP),
            (totales['otros_desc'],       FMT_CLP),
            (totales['total_descuentos'], FMT_CLP),
            (totales['sueldo_liquido'],   FMT_CLP),
        ]
        for col_idx, (valor, fmt) in enumerate(vals_total, start=1):
            c = ws.cell(row=fila_total, column=col_idx, value=valor)
            c.number_format = fmt
            c.fill      = fill_total
            c.border    = borde
            c.alignment = al_right if fmt == FMT_CLP else al_center
            if col_idx == 3:
                c.alignment = al_left
            c.font = ft_total_liq if col_idx == N_COLS else ft_total

        ws.page_setup.orientation   = 'landscape'
        ws.page_setup.paperSize     = ws.PAPERSIZE_LETTER
        ws.page_setup.fitToPage     = True
        ws.page_setup.fitToWidth    = 1
        ws.page_setup.fitToHeight   = 0
        ws.print_title_rows         = '1:5'
        ws.freeze_panes             = 'A6'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{nombre_base}.xlsx"'
        return response

    # ──────────────────────────────────────────────────────────────────────────
    @action(detail=False, methods=['get'], url_path='consolidado')
    def consolidado(self, request):
        anio_param = request.query_params.get('anio')
        mes_param  = request.query_params.get('mes')
        formato    = request.query_params.get('formato', 'json')

        if not anio_param:
            return Response({'error': 'Se requiere el parámetro anio.'}, status=400)
        try:
            anio = int(anio_param)
            mes  = int(mes_param) if mes_param else None
        except ValueError:
            return Response({'error': 'Los parámetros anio y mes deben ser numéricos.'}, status=400)

        base_qs = (
            Liquidacion.objects
            .filter(empleado__empresa__owner=request.user, anio=anio)
            .select_related('empleado', 'empleado__empresa', 'empleado__contrato_activo')
        )
        qs_periodo = base_qs.filter(mes=mes) if mes else base_qs

        if not qs_periodo.exists():
            if formato == 'json':
                return Response({'error': 'No hay liquidaciones para el período seleccionado.'}, status=404)
            return Response({'error': 'No hay liquidaciones para el período seleccionado.'}, status=404)

        def _costo_emp(liq):
            try:
                tipo = liq.empleado.contrato_activo.tipo_contrato
                tasa_afc = _TASA_AFC_EMP_INDEFINIDO if tipo == 'INDEFINIDO' else _TASA_AFC_EMP_PLAZO
            except Exception:
                tasa_afc = _TASA_AFC_EMP_INDEFINIDO
            return int(liq.total_imponible * (_TASA_SIS + _TASA_MUTUAL_AT + tasa_afc))

        def clp(n):
            if not n: return '$0'
            return f'${int(n):,}'.replace(',', '.')

        # ── Datos compartidos ─────────────────────────────────────────────────
        from collections import defaultdict
        masa_salarial = liquido_total = costo_empleador = 0
        empleados_ids = set()
        empresas_dict = defaultdict(lambda: {
            'id': None, 'nombre': '', 'rut': '',
            'trabajadores': 0, 'masa_salarial': 0,
            'liquido_total': 0, 'costo_empleador': 0,
        })
        for liq in qs_periodo:
            ce = _costo_emp(liq)
            masa_salarial   += liq.total_haberes
            liquido_total   += liq.sueldo_liquido
            costo_empleador += ce
            empleados_ids.add(liq.empleado_id)
            eid = liq.empleado.empresa_id
            emp = liq.empleado.empresa
            d = empresas_dict[eid]
            d['id']             = eid
            d['nombre']         = emp.nombre_legal
            d['rut']            = emp.rut
            d['trabajadores']  += 1
            d['masa_salarial'] += liq.total_haberes
            d['liquido_total'] += liq.sueldo_liquido
            d['costo_empleador'] += ce

        empresas_list = sorted(empresas_dict.values(), key=lambda x: x['masa_salarial'], reverse=True)

        MESES_CORTOS = ['Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
        MESES_LARGOS = ['Enero','Febrero','Marzo','Abril','Mayo','Junio',
                        'Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
        evolucion = []
        for m in range(1, 13):
            qs_m = base_qs.filter(mes=m)
            if qs_m.exists():
                ms = sum(l.total_haberes   for l in qs_m)
                lq = sum(l.sueldo_liquido  for l in qs_m)
                ce = sum(_costo_emp(l)     for l in qs_m)
                tw = qs_m.values('empleado_id').distinct().count()
            else:
                ms = lq = ce = tw = 0
            evolucion.append({'mes': m, 'mes_nombre': MESES_CORTOS[m-1],
                              'masa_salarial': ms, 'liquido_total': lq,
                              'costo_empleador': ce, 'trabajadores': tw})

        kpis = {
            'masa_salarial':   masa_salarial,
            'trabajadores':    len(empleados_ids),
            'costo_empleador': costo_empleador,
            'liquido_total':   liquido_total,
        }
        periodo_nombre = MESES_LARGOS[mes - 1] if mes else f'Año {anio}'
        nombre_base = f'Consolidado_{periodo_nombre.replace(" ","_")}_{anio}'

        # ══════════════════════════════════════════════════════════════════════
        # RAMA JSON
        # ══════════════════════════════════════════════════════════════════════
        if formato == 'json':
            return Response({
                'periodo': {'anio': anio, 'mes': mes, 'mes_nombre': MESES_LARGOS[mes-1] if mes else None},
                'kpis': kpis,
                'empresas': empresas_list,
                'evolucion': evolucion,
            })

        # ══════════════════════════════════════════════════════════════════════
        # RAMA EXCEL
        # ══════════════════════════════════════════════════════════════════════
        if formato == 'excel':
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()

            # ── Hoja 1: Resumen ───────────────────────────────────────────────
            ws = wb.active
            ws.title = 'Resumen'

            hdr_fill  = PatternFill('solid', fgColor='1E3A5F')
            tot_fill  = PatternFill('solid', fgColor='FEF3C7')
            kpi_fill  = PatternFill('solid', fgColor='0F2540')
            thin = Side(style='thin', color='CCCCCC')
            bord = Border(left=thin, right=thin, top=thin, bottom=thin)

            def _set(cell, val, bold=False, fill=None, align='left', color='000000'):
                cell.value = val
                cell.font  = Font(bold=bold, color=color, size=10)
                cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
                if fill: cell.fill = fill
                cell.border = bord

            # Title
            ws.merge_cells('A1:G1')
            t = ws['A1']
            t.value = f'Reporte Consolidado de Remuneraciones · {periodo_nombre} {anio}'
            t.font  = Font(bold=True, size=13, color='FFFFFF')
            t.fill  = PatternFill('solid', fgColor='0C1A35')
            t.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 28

            ws.merge_cells('A2:G2')
            ws['A2'].value = f'Generado el {timezone.now().date().strftime("%d/%m/%Y")} · {len(empresas_list)} empresa(s) · {kpis["trabajadores"]} trabajadore(s)'
            ws['A2'].font  = Font(size=9, color='888888')
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[2].height = 16

            # KPI row
            kpi_labels = ['Masa salarial', 'Trabajadores', 'Costo empleador', 'Líquido a pagar']
            kpi_vals   = [clp(kpis['masa_salarial']), str(kpis['trabajadores']),
                          clp(kpis['costo_empleador']), clp(kpis['liquido_total'])]
            kpi_cols   = [('A','B'), ('C','C'), ('D','E'), ('F','G')]
            ws.row_dimensions[3].height = 14
            ws.row_dimensions[4].height = 22
            ws.row_dimensions[5].height = 22
            ws.row_dimensions[6].height = 8
            for (c1, c2), lbl, val in zip(kpi_cols, kpi_labels, kpi_vals):
                ws.merge_cells(f'{c1}4:{c2}4')
                ws.merge_cells(f'{c1}5:{c2}5')
                lc = ws[f'{c1}4']
                lc.value = lbl; lc.font = Font(bold=True, size=8, color='AAAAAA')
                lc.fill = kpi_fill; lc.alignment = Alignment(horizontal='center', vertical='center')
                vc = ws[f'{c1}5']
                vc.value = val; vc.font = Font(bold=True, size=11, color='FFFFFF')
                vc.fill = kpi_fill; vc.alignment = Alignment(horizontal='center', vertical='center')

            # Table header
            cols_h = ['Empresa', 'RUT', 'Trabajadores', 'Masa Salarial', 'Costo Empleador', 'Líquido', '% del Total']
            for ci, h in enumerate(cols_h, 1):
                c = ws.cell(row=7, column=ci)
                _set(c, h, bold=True, fill=hdr_fill, align='center', color='FFFFFF')
            ws.row_dimensions[7].height = 20

            for ri, emp in enumerate(empresas_list, 8):
                pct = round(emp['masa_salarial'] / masa_salarial * 100, 1) if masa_salarial else 0
                row_fill = PatternFill('solid', fgColor='F0F4F8') if ri % 2 == 0 else None
                vals = [emp['nombre'], emp['rut'], emp['trabajadores'],
                        clp(emp['masa_salarial']), clp(emp['costo_empleador']),
                        clp(emp['liquido_total']), f'{pct}%']
                aligns = ['left','center','center','right','right','right','center']
                for ci, (v, a) in enumerate(zip(vals, aligns), 1):
                    _set(ws.cell(row=ri, column=ci), v, fill=row_fill, align=a)
                ws.row_dimensions[ri].height = 18

            # Totals
            tr = len(empresas_list) + 8
            tot_vals = ['TOTAL CONSOLIDADO', '', kpis['trabajadores'],
                        clp(kpis['masa_salarial']), clp(kpis['costo_empleador']),
                        clp(kpis['liquido_total']), '100%']
            for ci, v in enumerate(tot_vals, 1):
                a = 'right' if ci >= 4 else ('center' if ci == 3 else 'left')
                _set(ws.cell(row=tr, column=ci), v, bold=True, fill=tot_fill, align=a)
            ws.row_dimensions[tr].height = 20

            # Column widths
            for ci, w in enumerate([38, 14, 13, 18, 18, 18, 12], 1):
                ws.column_dimensions[get_column_letter(ci)].width = w
            ws.freeze_panes = 'A8'

            # ── Hoja 2: Evolución mensual ─────────────────────────────────────
            ws2 = wb.create_sheet('Evolución mensual')
            evo_headers = ['Mes', 'Masa Salarial', 'Líquido', 'Costo Empleador', 'Trabajadores']
            for ci, h in enumerate(evo_headers, 1):
                c = ws2.cell(row=1, column=ci)
                _set(c, h, bold=True, fill=hdr_fill, align='center', color='FFFFFF')
                ws2.row_dimensions[1].height = 20

            for ri, ev in enumerate(evolucion, 2):
                row_fill = PatternFill('solid', fgColor='F0F4F8') if ri % 2 == 0 else None
                vals = [ev['mes_nombre'], clp(ev['masa_salarial']), clp(ev['liquido_total']),
                        clp(ev['costo_empleador']), ev['trabajadores']]
                aligns = ['center','right','right','right','center']
                for ci, (v, a) in enumerate(zip(vals, aligns), 1):
                    _set(ws2.cell(row=ri, column=ci), v, fill=row_fill, align=a)
                ws2.row_dimensions[ri].height = 17

            for ci, w in enumerate([14, 18, 18, 18, 14], 1):
                ws2.column_dimensions[get_column_letter(ci)].width = w

            ws.page_setup.orientation = 'landscape'
            ws.page_setup.paperSize   = ws.PAPERSIZE_LETTER
            ws2.page_setup.orientation = 'landscape'
            ws2.page_setup.paperSize   = ws2.PAPERSIZE_LETTER

            buf = io.BytesIO()
            wb.save(buf); buf.seek(0)
            response = HttpResponse(buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{nombre_base}.xlsx"'
            return response

        # ══════════════════════════════════════════════════════════════════════
        # RAMA PDF
        # ══════════════════════════════════════════════════════════════════════
        if formato == 'pdf':
            for emp in empresas_list:
                emp['pct'] = round(emp['masa_salarial'] / masa_salarial * 100, 1) if masa_salarial else 0
                emp['masa_salarial_fmt']   = clp(emp['masa_salarial'])
                emp['costo_empleador_fmt'] = clp(emp['costo_empleador'])
                emp['liquido_total_fmt']   = clp(emp['liquido_total'])
            for ev in evolucion:
                ev['masa_salarial_fmt']   = clp(ev['masa_salarial'])
                ev['liquido_total_fmt']   = clp(ev['liquido_total'])
                ev['costo_empleador_fmt'] = clp(ev['costo_empleador'])

            context = {
                'periodo_nombre':  periodo_nombre,
                'anio':            anio,
                'mes':             mes,
                'kpis':            kpis,
                'kpis_fmt': {
                    'masa_salarial':   clp(kpis['masa_salarial']),
                    'costo_empleador': clp(kpis['costo_empleador']),
                    'liquido_total':   clp(kpis['liquido_total']),
                    'trabajadores':    kpis['trabajadores'],
                },
                'empresas':        empresas_list,
                'evolucion':       evolucion,
                'n_empresas':      len(empresas_list),
                'fecha_emision':   timezone.now().date().strftime('%d/%m/%Y'),
            }
            template = get_template('consolidado_remuneraciones.html')
            html = template.render(context)
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{nombre_base}.pdf"'
            pisa_status = pisa.CreatePDF(html, dest=response)
            if pisa_status.err:
                return Response({'error': 'Error al generar PDF'}, status=500)
            return response

        return Response({'error': 'Formato no válido. Use json, excel o pdf.'}, status=400)

# ==========================================
# PREVIRED EXPORT
# ==========================================

_AFP_CODIGOS_PREVIRED = {
    'CAPITAL': '03', 'CUPRUM': '04', 'HABITAT': '05',
    'MODELO': '08', 'PLANVITAL': '06', 'PROVIDA': '07', 'UNO': '10', 'IPS': '00',
}

_ISAPRE_CODIGOS_PREVIRED = {
    'BANMEDICA': '01', 'COLMENA': '02', 'CRUZ_BLANCA': '03',
    'ESENCIAL': '04', 'VIDA_TRES': '05', 'SAN_LORENZO': '06',
    'NUEVA_MASVIDA': '07', 'CONSALUD': '08',
}

_TIPO_CONTRATO_PREVIRED = {
    'INDEFINIDO': '1', 'PLAZO_FIJO': '2', 'OBRA_FAENA': '3',
}

_TIPO_JORNADA_PREVIRED = {
    'ORDINARIA': '1', 'TURNOS': '1', 'BISMANAL': '1',
    'ART_22': '2', 'PARCIAL': '2', 'OTRO': '1',
}

_MUTUAL_DEFAULT = '01'  # ISL por defecto
_TASA_MUTUAL_AT = 0.0093  # tasa base AT/EP empleador
_TASA_AFC_EMP_INDEFINIDO = 0.024
_TASA_AFC_EMP_PLAZO = 0.030
_TASA_SIS = 0.0149
_TASA_EXPECTATIVA_VIDA = 0.009  # Reforma 2025


def _rut_partes(rut_str: str):
    """Devuelve (numero_str, dv_str) desde un RUT como '12.345.678-9'."""
    limpio = (rut_str or '').replace('.', '').replace(' ', '').upper()
    if '-' in limpio:
        num, dv = limpio.rsplit('-', 1)
    elif len(limpio) > 1:
        num, dv = limpio[:-1], limpio[-1]
    else:
        return '0', '0'
    return num.lstrip('0') or '0', dv


def _fmt_fecha_previred(f) -> str:
    """Convierte fecha a DDMMAAAA o cadena vacía."""
    if not f:
        return ''
    try:
        if isinstance(f, str):
            d = datetime.date.fromisoformat(f)
        else:
            d = f
        return d.strftime('%d%m%Y')
    except Exception:
        return ''


# ==========================================
# FINIQUITO
# ==========================================

_TASAS_AFP = {
    'MODELO': 0.1058, 'HABITAT': 0.1127, 'PROVIDA': 0.1145,
    'CAPITAL': 0.1144, 'CUPRUM': 0.1144, 'PLANVITAL': 0.1116, 'UNO': 0.1049,
}
_VALOR_UF = 38000  # valor referencial

# Causales que dan derecho a indemnización por años (Art. 161)
_CAUSALES_CON_INDEMNIZACION = {'161_1', '161_2', '163bis'}


def _calcular_finiquito(empleado, fecha_termino, dias_trabajados_ultimo_mes, causal_articulo):
    """Precalcula todos los montos del finiquito a partir de datos del empleado."""
    contrato = Contrato.objects.filter(empleado=empleado).first()
    sueldo_base = contrato.sueldo_base if contrato else empleado.sueldo_base

    # Sueldo proporcional último mes
    sueldo_proporcional = math.floor((sueldo_base / 30) * dias_trabajados_ultimo_mes)

    # Gratificación proporcional (25% del sueldo proporcional, tope $200.000)
    gratificacion = min(math.floor(sueldo_proporcional * 0.25), 200_000)

    # Vacaciones adeudadas → feriado proporcional
    saldo_vac = calcular_saldo_vacaciones(empleado)
    dias_vac = saldo_vac['dias_disponibles']
    feriado_prop = math.floor((sueldo_base / 30) * dias_vac)

    # Indemnización por años de servicio (solo Art. 161 y 163bis)
    anos = (fecha_termino - empleado.fecha_ingreso).days // 365 if empleado.fecha_ingreso else 0
    if causal_articulo in _CAUSALES_CON_INDEMNIZACION:
        indemnizacion_anos = sueldo_base * min(anos, 11)
    else:
        indemnizacion_anos = 0

    # Descuentos previsionales sobre sueldo proporcional
    nombre_afp = (empleado.afp or 'MODELO').upper()
    tasa_afp = _TASAS_AFP.get(nombre_afp, 0.11)
    afp_monto = math.floor(sueldo_proporcional * tasa_afp)

    salud_nombre = (empleado.sistema_salud or 'FONASA').upper()
    if salud_nombre == 'ISAPRE' and empleado.plan_isapre_uf and float(empleado.plan_isapre_uf) > 0:
        salud_monto = max(math.floor(float(empleado.plan_isapre_uf) * _VALOR_UF),
                          math.floor(sueldo_proporcional * 0.07))
    else:
        salud_monto = math.floor(sueldo_proporcional * 0.07)

    descuentos_prevision = afp_monto + salud_monto

    total = (sueldo_proporcional + gratificacion + feriado_prop
             + indemnizacion_anos - descuentos_prevision)

    return {
        'sueldo_base':                  sueldo_base,
        'dias_trabajados_ultimo_mes':   dias_trabajados_ultimo_mes,
        'gratificacion_proporcional':   gratificacion,
        'feriado_proporcional':         feriado_prop,
        'indemnizacion_anos_servicio':  indemnizacion_anos,
        'indemnizacion_sustitutiva_aviso': 0,
        'otros_haberes':                0,
        'otros_descuentos':             0,
        'descuentos_prevision':         descuentos_prevision,
        'total_a_pagar':                max(total, 0),
    }


class FiniquitoViewSet(viewsets.ModelViewSet):
    serializer_class = FiniquitoSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = Finiquito.objects.filter(
            empleado__empresa__owner=self.request.user
        ).order_by('-fecha_emision')
        empleado_id = self.request.query_params.get('empleado')
        if empleado_id:
            qs = qs.filter(empleado_id=empleado_id)
        return qs

    def create(self, request, *args, **kwargs):
        if not _plan_permite(request.user, 2):
            return Response(
                {'error': 'Los finiquitos están disponibles desde el plan Starter. Mejora tu suscripción para acceder a esta función.'},
                status=status.HTTP_403_FORBIDDEN,
            )

        data = request.data
        empleado_id = data.get('empleado')

        try:
            empleado = Empleado.objects.get(id=empleado_id, empresa__owner=request.user)
        except Empleado.DoesNotExist:
            return Response({'error': 'Empleado no encontrado.'}, status=status.HTTP_404_NOT_FOUND)

        try:
            fecha_termino = datetime.date.fromisoformat(str(data.get('fecha_termino', '')))
        except (ValueError, TypeError):
            return Response({'error': 'Fecha de término inválida.'}, status=status.HTTP_400_BAD_REQUEST)

        dias = int(data.get('dias_trabajados_ultimo_mes', 30))
        causal = str(data.get('causal_articulo', ''))

        montos = _calcular_finiquito(empleado, fecha_termino, dias, causal)

        # Los montos pueden ser sobreescritos si el usuario los envía explícitamente
        for campo in montos:
            if campo in data and data[campo] is not None:
                montos[campo] = int(data[campo])

        # Recalcular total si algún monto fue sobreescrito
        montos['total_a_pagar'] = max(
            montos['sueldo_base'] // 30 * montos['dias_trabajados_ultimo_mes']
            + montos['gratificacion_proporcional']
            + montos['feriado_proporcional']
            + montos['indemnizacion_anos_servicio']
            + montos['indemnizacion_sustitutiva_aviso']
            + montos['otros_haberes']
            - montos['otros_descuentos']
            - montos['descuentos_prevision'],
            0,
        )

        finiquito = Finiquito.objects.create(
            empleado=empleado,
            documento_legal_id=data.get('documento_legal') or None,
            causal_articulo=causal,
            fecha_termino=fecha_termino,
            fecha_emision=datetime.date.fromisoformat(
                str(data.get('fecha_emision', datetime.date.today().isoformat()))
            ),
            modalidad=data.get('modalidad', 'PRESENCIAL'),
            **montos,
        )

        serializer = self.get_serializer(finiquito)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get'], url_path='generar_pdf')
    def generar_pdf(self, request, pk=None):
        try:
            finiquito = self.get_object()
            empleado  = finiquito.empleado
            empresa   = empleado.empresa

            def _fmt(f):
                if not f:
                    return '—'
                return f"{f.day:02d} de {_MESES[f.month - 1]} de {f.year}"

            ciudad = (getattr(empresa, 'comuna', '') or 'Santiago').strip().title()
            causal_label = finiquito.get_causal_articulo_display() if finiquito.causal_articulo else '—'

            sueldo_prop = math.floor(
                (finiquito.sueldo_base / 30) * finiquito.dias_trabajados_ultimo_mes
            )

            # Escapar campos de texto para prevenir inyección HTML/CSS en el PDF
            _ciudad      = _esc(ciudad)
            _causal      = _esc(causal_label)
            _nom_legal   = _esc(empresa.nombre_legal or '')
            _rut_emp     = _esc(empresa.rut or '')
            _trab_nombre = _esc(f"{empleado.nombres} {empleado.apellido_paterno} {empleado.apellido_materno or ''}")
            _trab_firma  = _esc(f"{empleado.nombres} {empleado.apellido_paterno}")
            _rut_trab    = _esc(empleado.rut or '')
            _cargo       = _esc(empleado.cargo or '—')
            _depto       = _esc(empleado.departamento or '—')
            _modalidad   = _esc(finiquito.get_modalidad_display())

            html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"/>
<style>
  @page {{ size: letter; margin: 2cm 2.5cm; }}
  body {{ font-family: Arial, sans-serif; font-size: 10pt; color: #111; line-height: 1.5; }}
  h1 {{ font-size: 14pt; text-align: center; text-transform: uppercase;
        letter-spacing: 2px; margin-bottom: 4px; }}
  h2 {{ font-size: 10pt; text-align: center; color: #555; margin-top: 0; margin-bottom: 20px; }}
  .seccion {{ margin-bottom: 16px; }}
  .seccion-titulo {{ font-size: 9pt; font-weight: bold; text-transform: uppercase;
                     letter-spacing: 1px; color: #555; border-bottom: 1px solid #ccc;
                     padding-bottom: 3px; margin-bottom: 8px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 10pt; }}
  table td {{ padding: 4px 6px; vertical-align: top; }}
  table td:last-child {{ text-align: right; font-weight: bold; }}
  .total-row td {{ border-top: 2px solid #333; font-weight: bold; font-size: 11pt;
                   padding-top: 8px; }}
  .firma-bloque {{ margin-top: 60px; display: flex; justify-content: space-between; }}
  .firma-item {{ text-align: center; width: 44%; }}
  .firma-linea {{ border-top: 1px solid #333; padding-top: 6px; margin-top: 50px; font-size: 9pt; }}
  p {{ margin: 4px 0; }}
  .aviso {{ font-size: 8pt; color: #666; margin-top: 20px; border-top: 1px solid #ccc; padding-top: 8px; }}
</style>
</head>
<body>

<h1>Finiquito de Contrato de Trabajo</h1>
<h2>{_ciudad}, {_fmt(finiquito.fecha_emision)}</h2>

<div class="seccion">
  <div class="seccion-titulo">Partes</div>
  <p><strong>Empleador:</strong> {_nom_legal} — RUT {_rut_emp}</p>
  <p><strong>Trabajador:</strong> {_trab_nombre} — RUT {_rut_trab}</p>
  <p><strong>Cargo:</strong> {_cargo} &nbsp;|&nbsp; <strong>Departamento:</strong> {_depto}</p>
  <p><strong>Fecha de ingreso:</strong> {_fmt(empleado.fecha_ingreso)} &nbsp;|&nbsp;
     <strong>Fecha de término:</strong> {_fmt(finiquito.fecha_termino)}</p>
  <p><strong>Causal de término:</strong> {_causal}</p>
</div>

<div class="seccion">
  <div class="seccion-titulo">Liquidación Final</div>
  <table>
    <tr><td>Sueldo base proporcional ({finiquito.dias_trabajados_ultimo_mes} días)</td>
        <td>${sueldo_prop:,.0f}</td></tr>
    <tr><td>Gratificación proporcional</td>
        <td>${finiquito.gratificacion_proporcional:,.0f}</td></tr>
    <tr><td>Feriado proporcional ({(finiquito.feriado_proporcional * 30 // finiquito.sueldo_base) if finiquito.sueldo_base else 0} días aprox.)</td>
        <td>${finiquito.feriado_proporcional:,.0f}</td></tr>
    {f'<tr><td>Indemnización por años de servicio (Art. 163)</td><td>${finiquito.indemnizacion_anos_servicio:,.0f}</td></tr>' if finiquito.indemnizacion_anos_servicio else ''}
    {f'<tr><td>Indemnización sustitutiva de aviso previo</td><td>${finiquito.indemnizacion_sustitutiva_aviso:,.0f}</td></tr>' if finiquito.indemnizacion_sustitutiva_aviso else ''}
    {f'<tr><td>Otros haberes</td><td>${finiquito.otros_haberes:,.0f}</td></tr>' if finiquito.otros_haberes else ''}
    <tr><td>Descuentos previsionales (AFP + Salud)</td>
        <td>-${finiquito.descuentos_prevision:,.0f}</td></tr>
    {f'<tr><td>Otros descuentos</td><td>-${finiquito.otros_descuentos:,.0f}</td></tr>' if finiquito.otros_descuentos else ''}
    <tr class="total-row">
      <td>TOTAL A PAGAR</td>
      <td>${finiquito.total_a_pagar:,.0f}</td>
    </tr>
  </table>
</div>

<div class="seccion">
  <div class="seccion-titulo">Declaración del Trabajador</div>
  <p>El trabajador declara haber recibido a su entera satisfacción la suma indicada como total a pagar,
  y nada más tiene que reclamar al empleador por concepto alguno derivado de la relación laboral
  que los vinculó, quedando ambas partes en paz y a finiquito.</p>
  <p>Modalidad de suscripción del finiquito: <strong>{_modalidad}</strong></p>
</div>

<div class="firma-bloque">
  <div class="firma-item">
    <div class="firma-linea">
      <strong>{_nom_legal}</strong><br/>RUT {_rut_emp}<br/>Empleador
    </div>
  </div>
  <div class="firma-item">
    <div class="firma-linea">
      <strong>{_trab_firma}</strong><br/>RUT {_rut_trab}<br/>Trabajador
    </div>
  </div>
</div>

<p class="aviso">
  Finiquito regulado por los artículos 177 y siguientes del Código del Trabajo de la República de Chile.
  Generado por Jornada40 · {_fmt(finiquito.fecha_emision)}.
</p>

</body>
</html>"""

            buffer = io.BytesIO()
            pisa_status = pisa.CreatePDF(html, dest=buffer)
            if pisa_status.err:
                return Response({'error': 'Error al generar el PDF.'}, status=500)

            buffer.seek(0)
            response = HttpResponse(buffer.read(), content_type='application/pdf')
            response['Content-Disposition'] = (
                f'attachment; filename="finiquito_{empleado.rut}_{finiquito.fecha_termino}.pdf"'
            )
            return response

        except Finiquito.DoesNotExist:
            return Response({'error': 'Finiquito no encontrado.'}, status=404)
        except Exception as e:
            return Response({'error': str(e)}, status=500)


# ==========================================
# FIRMA ELECTRÓNICA
# ==========================================

class SolicitudFirmaViewSet(viewsets.GenericViewSet):
    serializer_class = SolicitudFirmaSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return SolicitudFirma.objects.filter(
            empresa__owner=self.request.user
        ).select_related('empleado', 'empresa')

    def list(self, request):
        empleado_id = request.query_params.get('empleado_id')
        qs = self.get_queryset()
        if empleado_id:
            qs = qs.filter(empleado_id=empleado_id)
        return Response(self.get_serializer(qs, many=True).data)

    @action(detail=False, methods=['post'])
    def solicitar(self, request):
        empleado_id   = request.data.get('empleado_id')
        tipo_doc      = request.data.get('tipo_documento')
        contrato_id   = request.data.get('contrato_id')
        doc_legal_id  = request.data.get('documento_legal_id')
        anexo_id      = request.data.get('anexo_contrato_id')
        liquidacion_id = request.data.get('liquidacion_id')
        vacacion_id    = request.data.get('vacacion_id')
        finiquito_id   = request.data.get('finiquito_id')

        tipos_validos = [t[0] for t in SolicitudFirma.TIPOS_DOCUMENTO]
        if tipo_doc not in tipos_validos:
            return Response({'error': 'Tipo de documento inválido.'}, status=400)

        try:
            empleado = Empleado.objects.get(id=empleado_id, empresa__owner=request.user)
        except Empleado.DoesNotExist:
            return Response({'error': 'Trabajador no encontrado.'}, status=404)

        empresa = empleado.empresa

        if not empresa.firma_imagen:
            return Response(
                {'error': 'La empresa no tiene firma del empleador configurada. Configúrela en el Lobby de Empresas.'},
                status=400
            )

        email_trabajador = empleado.email
        if not email_trabajador:
            return Response(
                {'error': 'El trabajador no tiene email registrado. Agréguelo en Datos Generales antes de enviar a firma.'},
                status=400
            )

        es_plan_semilla = _es_plan_semilla(request.user)

        try:
            pdf_bytes, contrato_obj, doc_legal_obj, liquidacion_obj, vacacion_obj, finiquito_obj = self._generar_pdf_firma(
                empleado, empresa, tipo_doc,
                contrato_id, doc_legal_id, anexo_id,
                liquidacion_id, vacacion_id, finiquito_id, es_plan_semilla
            )
        except Exception as e:
            return Response({'error': str(e)}, status=400)

        key = b2_client.key_pendiente(empresa.id, str(uuid_mod.uuid4()))
        try:
            b2_client.subir_documento(pdf_bytes, key)
        except RuntimeError as e:
            return Response({'error': str(e)}, status=503)
        except Exception:
            return Response({'error': 'Error al subir el documento al almacenamiento.'}, status=500)

        try:
            solicitud = SolicitudFirma.objects.create(
                empleado=empleado,
                empresa=empresa,
                tipo_documento=tipo_doc,
                contrato=contrato_obj,
                documento_legal=doc_legal_obj,
                liquidacion=liquidacion_obj,
                vacacion=vacacion_obj,
                finiquito=finiquito_obj,
                email_firmante=email_trabajador,
                b2_key_temporal=key,
            )
        except Exception as e:
            b2_client.eliminar_documento(key)
            return Response({'error': f'Error al registrar la solicitud: {e}'}, status=500)

        try:
            self._enviar_email_firma(solicitud, empleado, empresa)
        except Exception:
            pass

        return Response(SolicitudFirmaSerializer(solicitud).data, status=201)

    @action(detail=True, methods=['patch'])
    def cancelar(self, request, pk=None):
        solicitud = self.get_object()
        if solicitud.estado != 'PENDIENTE':
            return Response({'error': 'Solo se puede cancelar una solicitud pendiente.'}, status=400)
        solicitud.estado = 'CANCELADO'
        solicitud.save(update_fields=['estado', 'actualizado_en'])
        return Response(SolicitudFirmaSerializer(solicitud).data)

    @action(detail=True, methods=['post'])
    def reenviar(self, request, pk=None):
        solicitud = self.get_object()
        if solicitud.estado != 'PENDIENTE':
            return Response({'error': 'Solo se puede reenviar una solicitud pendiente.'}, status=400)
        try:
            self._enviar_email_firma(solicitud, solicitud.empleado, solicitud.empresa)
        except Exception as e:
            return Response({'error': f'Error al reenviar el email: {str(e)}'}, status=500)
        return Response({'mensaje': 'Email de firma reenviado correctamente.'})

    @action(detail=True, methods=['post'])
    def descargar(self, request, pk=None):
        """Genera una URL presignada de corta duración para descargar el PDF firmado."""
        solicitud = self.get_object()
        if solicitud.estado != 'FIRMADO' or not solicitud.b2_key_firmado:
            return Response({'error': 'No hay PDF firmado disponible para esta solicitud.'}, status=400)
        try:
            url = b2_client.generar_url_presignada(solicitud.b2_key_firmado, ttl_segundos=300)
        except RuntimeError as exc:
            return Response({'error': str(exc)}, status=503)
        except Exception:
            return Response({'error': 'Error al generar el enlace de descarga.'}, status=500)
        return Response({'url': url})

    # -------------------------------------------------
    # Helpers internos
    # -------------------------------------------------

    MESES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
             "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]

    def _generar_pdf_firma(self, empleado, empresa, tipo_doc,
                           contrato_id, doc_legal_id, anexo_id,
                           liquidacion_id, vacacion_id, finiquito_id, es_plan_semilla):
        """Genera el PDF a firmar y retorna (pdf_bytes, contrato, doc_legal, liquidacion, vacacion, finiquito)."""
        ciudad = str(
            getattr(empresa, 'ciudad', '') or getattr(empresa, 'comuna', '') or 'Santiago'
        ).strip().title()

        if tipo_doc == 'CONTRATO':
            try:
                contrato = (Contrato.objects.get(id=contrato_id)
                            if contrato_id else Contrato.objects.get(empleado=empleado))
            except Contrato.DoesNotExist:
                raise Exception('El trabajador no tiene contrato registrado.')
            ctx = _ctx_contrato(contrato, es_plan_semilla)
            html = render_to_string('contrato_trabajo.html', ctx)
            return _html_a_pdf_bytes(html, f'Contrato_{empleado.rut}'), contrato, None, None, None, None

        if tipo_doc == 'ANEXO_40H':
            try:
                contrato = (Contrato.objects.get(id=contrato_id)
                            if contrato_id else Contrato.objects.get(empleado=empleado))
            except Contrato.DoesNotExist:
                raise Exception('El trabajador no tiene contrato registrado.')
            ctx = _ctx_contrato(contrato, es_plan_semilla)
            html = render_to_string('anexo_40h.html', ctx)
            return _html_a_pdf_bytes(html, f'Anexo40h_{empleado.rut}'), contrato, None, None, None, None

        if tipo_doc in ('AMONESTACION', 'CONSTANCIA'):
            if doc_legal_id:
                try:
                    doc = DocumentoLegal.objects.get(id=doc_legal_id, empleado=empleado)
                except DocumentoLegal.DoesNotExist:
                    raise Exception('Documento legal no encontrado.')
            else:
                doc = DocumentoLegal.objects.filter(
                    empleado=empleado, tipo=tipo_doc
                ).order_by('-fecha_emision').first()
                if not doc:
                    raise Exception(f'No se encontró documento de tipo {tipo_doc}.')
            hoy = doc.fecha_emision
            fecha_es = f"{hoy.day:02d} de {self.MESES[hoy.month - 1]} de {hoy.year}"
            ctx = {'documento': doc, 'empleado': empleado, 'empresa': empresa,
                   'fecha_actual': fecha_es, 'ciudad': ciudad,
                   'es_plan_semilla': es_plan_semilla}
            html = render_to_string('documento_legal.html', ctx)
            return _html_a_pdf_bytes(html, f'{doc.tipo}_{empleado.rut}'), None, doc, None, None, None

        if tipo_doc == 'DESPIDO':
            if doc_legal_id:
                try:
                    doc = DocumentoLegal.objects.get(id=doc_legal_id, empleado=empleado)
                except DocumentoLegal.DoesNotExist:
                    raise Exception('Documento legal no encontrado.')
            else:
                doc = DocumentoLegal.objects.filter(
                    empleado=empleado, tipo='DESPIDO'
                ).order_by('-fecha_emision').first()
                if not doc:
                    raise Exception('No se encontró carta de despido.')
            pdf_bytes = self._pdf_para_documento_legal(doc, es_plan_semilla)
            return pdf_bytes, None, doc, None, None, None

        if tipo_doc == 'ANEXO_CONTRATO':
            if not anexo_id:
                raise Exception('Se requiere el ID del anexo de contrato.')
            try:
                contrato = Contrato.objects.get(empleado=empleado)
                anexo = AnexoContrato.objects.get(id=anexo_id, contrato=contrato)
            except (Contrato.DoesNotExist, AnexoContrato.DoesNotExist):
                raise Exception('Anexo de contrato no encontrado.')
            hoy = anexo.fecha_emision
            fecha_es = f"{hoy.day:02d} de {self.MESES[hoy.month - 1]} de {hoy.year}"
            ctx = {'anexo': anexo, 'contrato': contrato, 'empleado': empleado,
                   'empresa': empresa, 'fecha_actual': fecha_es, 'ciudad': ciudad,
                   'es_plan_semilla': es_plan_semilla}
            html = render_to_string('anexo_contrato.html', ctx)
            return _html_a_pdf_bytes(html, f'AnexoContrato_{empleado.rut}'), contrato, None, None, None, None

        if tipo_doc == 'LIQUIDACION':
            if not liquidacion_id:
                raise Exception('Se requiere el ID de la liquidación.')
            try:
                liq = Liquidacion.objects.get(id=liquidacion_id, empleado=empleado)
            except Liquidacion.DoesNotExist:
                raise Exception('Liquidación no encontrada.')
            contrato_liq = Contrato.objects.filter(empleado=empleado).first()
            meses = self.MESES
            mes_nombre = meses[liq.mes - 1]
            sueldo_seguro = int(liq.sueldo_liquido or 0)
            liquido_palabras = num2words(sueldo_seguro, lang='es')
            det_no_imp = liq.detalle_haberes_no_imponibles
            if not isinstance(det_no_imp, list): det_no_imp = []
            suma_no_imponibles = sum(int(i.get('valor', 0)) for i in det_no_imp if isinstance(i, dict))
            det_otros = liq.detalle_otros_descuentos
            if not isinstance(det_otros, list): det_otros = []
            suma_otros_descuentos = sum(int(i.get('valor', 0)) for i in det_otros if isinstance(i, dict))
            total_ley = ((liq.afp_monto or 0) + (liq.salud_monto or 0) +
                         (liq.seguro_cesantia or 0) + (liq.impuesto_unico or 0))
            total_otros_dsctos = (liq.anticipo_quincena or 0) + suma_otros_descuentos
            ctx = {
                'liquidacion': liq, 'empleado': empleado, 'empresa': empresa,
                'contrato': contrato_liq,
                'mes_nombre': mes_nombre.upper(),
                'liquido_palabras': liquido_palabras,
                'total_no_imponible': suma_no_imponibles,
                'total_ley': total_ley,
                'total_otros_dsctos': total_otros_dsctos,
                'es_plan_semilla': es_plan_semilla,
            }
            html = render_to_string('liquidacion.html', ctx)
            return _html_a_pdf_bytes(html, f'Liquidacion_{liq.mes}_{liq.anio}_{empleado.rut}'), None, None, liq, None, None

        if tipo_doc == 'VACACION':
            if not vacacion_id:
                raise Exception('Se requiere el ID del comprobante de vacaciones.')
            try:
                vac = VacacionEmpleado.objects.get(id=vacacion_id, empleado=empleado)
            except VacacionEmpleado.DoesNotExist:
                raise Exception('Comprobante de vacaciones no encontrado.')
            hoy = datetime.date.today()
            fecha_hoy = f"{hoy.day:02d} de {self.MESES[hoy.month - 1]} de {hoy.year}"
            def _fmt(f):
                return f"{f.day:02d} de {self.MESES[f.month - 1]} de {f.year}" if f else '—'
            ctx = {
                'vacacion': vac, 'empleado': empleado, 'empresa': empresa,
                'fecha_actual': fecha_hoy,
                'fecha_inicio_texto': _fmt(vac.fecha_inicio),
                'fecha_fin_texto': _fmt(vac.fecha_fin),
                'ciudad': ciudad,
                'es_plan_semilla': es_plan_semilla,
            }
            html = render_to_string('comprobante_vacaciones.html', ctx)
            return _html_a_pdf_bytes(html, f'Vacacion_{empleado.rut}'), None, None, None, vac, None

        if tipo_doc == 'FINIQUITO':
            if not finiquito_id:
                raise Exception('Se requiere el ID del finiquito.')
            try:
                fin = Finiquito.objects.get(id=finiquito_id, empleado=empleado)
            except Finiquito.DoesNotExist:
                raise Exception('Finiquito no encontrado.')

            def _fmt_fin(f):
                if not f:
                    return '—'
                return f"{f.day:02d} de {_MESES[f.month - 1]} de {f.year}"

            ciudad_fin = (getattr(empresa, 'ciudad', '') or getattr(empresa, 'comuna', '') or 'Santiago').strip().title()
            causal_label = fin.get_causal_articulo_display() if fin.causal_articulo else '—'
            sueldo_prop = math.floor((fin.sueldo_base / 30) * fin.dias_trabajados_ultimo_mes)
            _f_ciudad     = _esc(ciudad_fin)
            _f_causal     = _esc(causal_label)
            _f_nom_legal  = _esc(empresa.nombre_legal or '')
            _f_rut_emp    = _esc(empresa.rut or '')
            _f_trab_nomb  = _esc(f"{empleado.nombres} {empleado.apellido_paterno} {empleado.apellido_materno or ''}")
            _f_trab_firma = _esc(f"{empleado.nombres} {empleado.apellido_paterno}")
            _f_rut_trab   = _esc(empleado.rut or '')
            _f_cargo      = _esc(empleado.cargo or '—')
            _f_depto      = _esc(empleado.departamento or '—')
            _f_modalidad  = _esc(fin.get_modalidad_display())
            html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8"/>
<style>
  @page {{ size: letter; margin: 2cm 2.5cm; }}
  body {{ font-family: Arial, sans-serif; font-size: 10pt; color: #111; line-height: 1.5; }}
  h1 {{ font-size: 14pt; text-align: center; text-transform: uppercase; letter-spacing: 2px; margin-bottom: 4px; }}
  h2 {{ font-size: 10pt; text-align: center; color: #555; margin-top: 0; margin-bottom: 20px; }}
  .seccion {{ margin-bottom: 16px; }}
  .seccion-titulo {{ font-size: 9pt; font-weight: bold; text-transform: uppercase;
                     letter-spacing: 1px; color: #555; border-bottom: 1px solid #ccc;
                     padding-bottom: 3px; margin-bottom: 8px; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 10pt; }}
  table td {{ padding: 4px 6px; vertical-align: top; }}
  table td:last-child {{ text-align: right; font-weight: bold; }}
  .total-row td {{ border-top: 2px solid #333; font-weight: bold; font-size: 11pt; padding-top: 8px; }}
  .firma-bloque {{ margin-top: 60px; display: flex; justify-content: space-between; }}
  .firma-item {{ text-align: center; width: 44%; }}
  .firma-linea {{ border-top: 1px solid #333; padding-top: 6px; margin-top: 50px; font-size: 9pt; }}
  p {{ margin: 4px 0; }}
  .aviso {{ font-size: 8pt; color: #666; margin-top: 20px; border-top: 1px solid #ccc; padding-top: 8px; }}
</style>
</head>
<body>
<h1>Finiquito de Contrato de Trabajo</h1>
<h2>{_f_ciudad}, {_fmt_fin(fin.fecha_emision)}</h2>
<div class="seccion">
  <div class="seccion-titulo">Partes</div>
  <p><strong>Empleador:</strong> {_f_nom_legal} — RUT {_f_rut_emp}</p>
  <p><strong>Trabajador:</strong> {_f_trab_nomb} — RUT {_f_rut_trab}</p>
  <p><strong>Cargo:</strong> {_f_cargo} &nbsp;|&nbsp; <strong>Departamento:</strong> {_f_depto}</p>
  <p><strong>Fecha de ingreso:</strong> {_fmt_fin(empleado.fecha_ingreso)} &nbsp;|&nbsp;
     <strong>Fecha de término:</strong> {_fmt_fin(fin.fecha_termino)}</p>
  <p><strong>Causal de término:</strong> {_f_causal}</p>
</div>
<div class="seccion">
  <div class="seccion-titulo">Liquidación Final</div>
  <table>
    <tr><td>Sueldo base proporcional ({fin.dias_trabajados_ultimo_mes} días)</td><td>${sueldo_prop:,.0f}</td></tr>
    <tr><td>Gratificación proporcional</td><td>${fin.gratificacion_proporcional:,.0f}</td></tr>
    <tr><td>Feriado proporcional</td><td>${fin.feriado_proporcional:,.0f}</td></tr>
    {f'<tr><td>Indemnización por años de servicio</td><td>${fin.indemnizacion_anos_servicio:,.0f}</td></tr>' if fin.indemnizacion_anos_servicio else ''}
    {f'<tr><td>Indemnización sustitutiva de aviso previo</td><td>${fin.indemnizacion_sustitutiva_aviso:,.0f}</td></tr>' if fin.indemnizacion_sustitutiva_aviso else ''}
    {f'<tr><td>Otros haberes</td><td>${fin.otros_haberes:,.0f}</td></tr>' if fin.otros_haberes else ''}
    <tr><td>Descuentos previsionales</td><td>-${fin.descuentos_prevision:,.0f}</td></tr>
    {f'<tr><td>Otros descuentos</td><td>-${fin.otros_descuentos:,.0f}</td></tr>' if fin.otros_descuentos else ''}
    <tr class="total-row"><td>TOTAL A PAGAR</td><td>${fin.total_a_pagar:,.0f}</td></tr>
  </table>
</div>
<div class="seccion">
  <div class="seccion-titulo">Declaración del Trabajador</div>
  <p>El trabajador declara haber recibido a su entera satisfacción la suma indicada como total a pagar,
  y nada más tiene que reclamar al empleador, quedando ambas partes en paz y a finiquito.</p>
  <p>Modalidad de suscripción: <strong>{_f_modalidad}</strong></p>
</div>
<div class="firma-bloque">
  <div class="firma-item">
    <div class="firma-linea"><strong>{_f_nom_legal}</strong><br/>RUT {_f_rut_emp}<br/>Empleador</div>
  </div>
  <div class="firma-item">
    <div class="firma-linea"><strong>{_f_trab_firma}</strong><br/>RUT {_f_rut_trab}<br/>Trabajador</div>
  </div>
</div>
<p class="aviso">Finiquito regulado por los artículos 177 y siguientes del Código del Trabajo de Chile.
  Generado por Jornada40 · {_fmt_fin(fin.fecha_emision)}.</p>
</body>
</html>"""
            return _html_a_pdf_bytes(html, f'Finiquito_{empleado.rut}'), None, None, None, None, fin

        raise Exception(f'Tipo de documento no soportado: {tipo_doc}')

    def _enviar_email_firma(self, solicitud, empleado, empresa):
        tipo_labels = {
            'CONTRATO':       'Contrato Laboral',
            'ANEXO_40H':      'Anexo Ley 40 Horas',
            'AMONESTACION':   'Carta de Amonestación',
            'DESPIDO':        'Carta de Despido',
            'CONSTANCIA':     'Constancia Laboral',
            'ANEXO_CONTRATO': 'Anexo de Contrato',
            'LIQUIDACION':    'Liquidación de Sueldo',
            'VACACION':       'Comprobante de Vacaciones',
            'FINIQUITO':      'Finiquito de Término',
        }
        tipo_label       = tipo_labels.get(solicitud.tipo_documento, solicitud.tipo_documento)
        firma_url        = f"https://jornada40.cl/firma/{solicitud.token}"
        nombre_trabajador = f"{empleado.nombres} {empleado.apellido_paterno}"
        expira_fecha     = solicitud.expira_en.strftime('%d/%m/%Y')

        texto_plano = (
            f"Hola {nombre_trabajador},\n\n"
            f"{empresa.nombre_legal} requiere tu firma en: {tipo_label}.\n\n"
            f"Para firmar ingresa al siguiente enlace (válido hasta el {expira_fecha}):\n"
            f"{firma_url}\n\n"
            f"Si no reconoces esta solicitud, ignora este mensaje.\n\n"
            f"Jornada40 — Sistema de Gestión Laboral"
        )
        html_body = f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;background:#f4f6f9;margin:0;padding:0;">
  <div style="max-width:560px;margin:40px auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.08);">
    <div style="background:linear-gradient(135deg,#0c1a35,#1e3a6e);padding:32px 40px;">
      <h1 style="color:#fff;margin:0;font-size:22px;font-weight:700;">Firma Electrónica Requerida</h1>
      <p style="color:rgba(255,255,255,0.6);margin:6px 0 0;font-size:14px;">{empresa.nombre_legal}</p>
    </div>
    <div style="padding:32px 40px;">
      <p style="color:#374151;font-size:15px;margin:0 0 8px;">Hola <strong>{nombre_trabajador}</strong>,</p>
      <p style="color:#374151;font-size:14px;line-height:1.6;margin:0 0 24px;">
        Tu empleador requiere tu firma electrónica en el siguiente documento:
      </p>
      <div style="background:#f0f4ff;border-left:4px solid #2563eb;padding:16px 20px;border-radius:6px;margin-bottom:28px;">
        <p style="margin:0;font-weight:700;color:#1e3a6e;font-size:15px;">{tipo_label}</p>
        <p style="margin:4px 0 0;color:#6b7280;font-size:13px;">Válido para firmar hasta el {expira_fecha}</p>
      </div>
      <a href="{firma_url}"
         style="display:inline-block;background:linear-gradient(135deg,#2563eb,#1d4ed8);color:#fff;font-weight:700;font-size:15px;padding:14px 32px;border-radius:8px;text-decoration:none;">
        Revisar y Firmar Documento
      </a>
      <p style="color:#9ca3af;font-size:12px;margin:28px 0 0;line-height:1.6;">
        Si el botón no funciona, copia este enlace:<br>
        <a href="{firma_url}" style="color:#2563eb;word-break:break-all;">{firma_url}</a>
      </p>
      <hr style="border:none;border-top:1px solid #e5e7eb;margin:24px 0;">
      <p style="color:#9ca3af;font-size:11px;margin:0;line-height:1.6;">
        Mensaje enviado a trabajador de <strong>{empresa.nombre_legal}</strong>.
        Firma Electrónica Simple válida bajo Ley 19.799 (Chile).
      </p>
    </div>
  </div>
</body>
</html>"""
        msg = EmailMultiAlternatives(
            subject=f"Firma requerida: {tipo_label} — {empresa.nombre_legal}",
            body=texto_plano,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[solicitud.email_firmante],
        )
        msg.attach_alternative(html_body, "text/html")
        msg.send()


# Endpoint para listar los planes activos en la BD
class PlanViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Plan.objects.filter(activo=True)
    serializer_class = PlanSerializer
    permission_classes = [AllowAny] 

# Endpoint específico para el dashboard del cliente
@api_view(['GET'])
@permission_classes([IsAuthenticated])
def mi_suscripcion(request):
    cliente = getattr(request.user, 'perfil_cliente', None)
    if not cliente:
        return Response({'error': 'Perfil de cliente no encontrado'}, status=status.HTTP_404_NOT_FOUND)

    # 1. Buscar o crear suscripción 
    try:
        suscripcion = cliente.suscripcion_activa
    except Suscripcion.DoesNotExist:
        plan_asignado = cliente.plan if cliente.plan else Plan.objects.first()
        suscripcion = Suscripcion.objects.create(
            cliente=cliente,
            plan=plan_asignado,
            estado='ACTIVE' if plan_asignado else 'TRIAL'
        )

    # 2. Calcular uso real (trabajadores en TODAS las empresas del usuario)
    trabajadores_actuales = Empleado.objects.filter(empresa__owner=request.user).count()

    # 3. Armar la respuesta exacta que espera Suscripcion.tsx
    data = {
        'estado': suscripcion.estado,
        'plan': {
            'id': suscripcion.plan.id,
            'nombre': suscripcion.plan.nombre,
            'precio': suscripcion.plan.precio,
            'limite_trabajadores': suscripcion.plan.limite_trabajadores,
            'descripcion': suscripcion.plan.descripcion,
        },
        'trabajadores_actuales': trabajadores_actuales,
        'fecha_proximo_cobro': suscripcion.fecha_proximo_cobro.strftime('%Y-%m-%d') if suscripcion.fecha_proximo_cobro else None,
        'metodo_pago_glosa': suscripcion.metodo_pago_glosa,
    }

    return Response(data, status=status.HTTP_200_OK)

# ==========================================
# PASARELA DE PAGOS (REVENIU)
# ==========================================

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def crear_checkout_reveniu(request):
    plan_id = str(request.data.get('plan_id'))
    ciclo = request.data.get('ciclo', 'mensual')  # 'mensual' o 'anual'

    try:
        plan = Plan.objects.get(id=plan_id)
        cliente = getattr(request.user, 'perfil_cliente', None)

        nombre_plan = plan.nombre.upper()
        ciclo_upper = ciclo.upper()
        env_key = f'REVENIU_LINK_{nombre_plan}_{ciclo_upper}'
        link_base = config(env_key, default=None)

        if not link_base:
            return Response(
                {'error': f'Link de pago no configurado para este plan ({env_key} no definido).'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        nombre_url = urllib.parse.quote(f"{request.user.first_name} {request.user.last_name}".strip())
        url_pago = f"{link_base}?email={request.user.email}&name={nombre_url}&custom_reference={cliente.id}_{plan.id}"

        return Response({'url': url_pago}, status=status.HTTP_200_OK)

    except Plan.DoesNotExist:
        return Response({'error': 'Plan no encontrado.'}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
@permission_classes([AllowAny])
def webhook_reveniu(request):
    webhook_secret = config('REVENIU_WEBHOOK_SECRET', default=None)

    # Secret obligatorio. Sin él, rechazamos todo (fail closed).
    if not webhook_secret:
        return Response({'error': 'Webhook no configurado'}, status=status.HTTP_503_SERVICE_UNAVAILABLE)

    token_recibido = request.headers.get('X-Webhook-Token', '')

    if not hmac.compare_digest(token_recibido, webhook_secret):
        return Response({'error': 'Unauthorized'}, status=status.HTTP_401_UNAUTHORIZED)

    data = request.data
    evento = data.get('event')

    if evento in ['subscription_created', 'payment_succeeded']:
        custom_reference = data.get('custom_reference', '')

        if not custom_reference or '_' not in custom_reference:
            return Response({'error': 'custom_reference inválido'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            cliente_id, plan_id = custom_reference.split('_', 1)
            cliente_id = int(cliente_id)
            plan_id = int(plan_id)
        except (ValueError, AttributeError):
            return Response({'error': 'Formato de custom_reference inválido'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            suscripcion = Suscripcion.objects.get(cliente_id=cliente_id)
            plan_nuevo = Plan.objects.get(id=plan_id)
        except (Suscripcion.DoesNotExist, Plan.DoesNotExist):
            return Response({'error': 'Suscripción o plan no encontrado'}, status=status.HTTP_400_BAD_REQUEST)

        suscripcion.plan = plan_nuevo
        suscripcion.estado = 'ACTIVE'
        suscripcion.gateway_subscription_id = str(data.get('subscription_id', ''))
        suscripcion.save()

        # Sincronizar también cliente.plan para que los PDFs y límites
        # reflejen el plan activo sin depender de la Suscripcion.
        suscripcion.cliente.plan = plan_nuevo
        suscripcion.cliente.save(update_fields=['plan'])

    return Response(status=status.HTTP_200_OK)

@api_view(['POST'])
@permission_classes([AllowAny])
@throttle_classes([PasswordResetRateThrottle])
def recuperar_password_por_rut(request):
    rut = request.data.get('rut')
    
    if not rut:
        return Response({'error': 'Debes ingresar un RUT'}, status=400)

    try:
        # 1. VALIDAMOS CONTRA TABLA CLIENTES
        cliente = Cliente.objects.get(rut=rut)
    except Cliente.DoesNotExist:
        # Mensaje ambiguo por seguridad anti-hackers
        return Response({'error': 'Si el RUT es válido, enviaremos un correo.'}, status=404)

    # 2. VERIFICAMOS SI EL CLIENTE TIENE CORREO
    if not cliente.correo:
        return Response({'error': 'Si el RUT es válido, enviaremos un correo.'}, status=404)

    try:
        # 3. BUSCAMOS AL USUARIO DE DJANGO ASOCIADO A ESE RUT
        user = User.objects.get(username=rut)
    except User.DoesNotExist:
        return Response({'error': 'Error interno: Cliente existe pero no tiene credenciales de acceso.'}, status=400)

    if user.email != cliente.correo:
        user.email = cliente.correo
        user.save()

    # 4. ENMASCARAR EL CORREO (con***@gmail.com)
    partes = cliente.correo.split('@')
    if len(partes) != 2:
        return Response({'error': 'El correo registrado tiene un formato inválido.'}, status=400)
    nombre_correo, dominio = partes

    if len(nombre_correo) > 3:
        correo_oculto = nombre_correo[:3] + '*' * (len(nombre_correo) - 3) + '@' + dominio
    elif len(nombre_correo) > 1:
        correo_oculto = nombre_correo[0] + '*' * (len(nombre_correo) - 1) + '@' + dominio
    else:
        correo_oculto = '*@' + dominio

    # 5. ENVIAR EL CORREO
    form = PasswordResetForm({'email': user.email})
    if form.is_valid():
        form.save(
            request=request,
            use_https=True,
            from_email=settings.DEFAULT_FROM_EMAIL,
            email_template_name='registration/password_reset_email.html',
            html_email_template_name='registration/password_reset_email.html',
        )

    # 6. RESPONDER A REACT
    return Response({
        'mensaje': 'Correo enviado exitosamente', 
        'correo_oculto': correo_oculto
    }, status=200)

@api_view(['GET', 'PUT'])
@permission_classes([IsAuthenticated])
def perfil_usuario(request):
    # Obtenemos el perfil del cliente (Tabla Cliente)
    cliente = getattr(request.user, 'perfil_cliente', None)
    
    if not cliente:
        return Response({'error': 'Perfil no encontrado en la tabla Cliente'}, status=404)

    if request.method == 'GET':
        # EXTRAEMOS DIRECTAMENTE DE LA TABLA CLIENTE
        return Response({
            'nombres': cliente.nombres or "",
            'apellido_paterno': cliente.apellido_paterno or "",
            'apellido_materno': cliente.apellido_materno or "",
            'email': request.user.email,
            'rut': cliente.rut,
            # Agregamos el nombre del plan para verificar el error del null
            'plan_nombre': cliente.plan.nombre if cliente.plan else "Sin Plan (null)"
        })

    if request.method == 'PUT':
        nombres = request.data.get('nombres')
        paterno = request.data.get('apellido_paterno')
        materno = request.data.get('apellido_materno')
        email = request.data.get('email')

        # 1. ACTUALIZAR TABLA CLIENTE (Campos específicos)
        if nombres is not None: cliente.nombres = nombres
        if paterno is not None: cliente.apellido_paterno = paterno
        if materno is not None: cliente.apellido_materno = materno
        if email: cliente.correo = email
        cliente.save()

        # 2. MANTENER REDUNDANCIA EN TABLA USER (Django Auth)
        if nombres is not None: request.user.first_name = nombres
        request.user.last_name = f"{paterno or ''} {materno or ''}".strip()
        if email: request.user.email = email
        request.user.save()

        return Response({'mensaje': 'Perfil actualizado con éxito en ambas tablas'})


# ==========================================
# FIRMA ELECTRÓNICA — ENDPOINTS PÚBLICOS
# (no requieren autenticación del empleador)
# ==========================================

def _enmascarar_email(email: str) -> str:
    """Ej: juan.perez@gmail.com → ju***@gmail.com"""
    partes = email.split('@')
    usuario = partes[0]
    dominio = partes[1] if len(partes) > 1 else ''
    prefijo = usuario[:2] if len(usuario) >= 2 else usuario[:1]
    return f"{prefijo}***@{dominio}"


@api_view(['GET'])
@permission_classes([AllowAny])
def firma_publica_info(request, token):
    """
    Retorna la información pública de una solicitud de firma:
    tipo de documento, nombre de empresa, nombre del trabajador y fecha de expiración.
    No expone datos sensibles.
    """
    try:
        solicitud = SolicitudFirma.objects.select_related('empleado', 'empresa').get(token=token)
    except SolicitudFirma.DoesNotExist:
        return Response({'error': 'Solicitud de firma no encontrada.'}, status=404)

    # Marcar como expirada si corresponde
    if solicitud.estado == 'PENDIENTE' and timezone.now() > solicitud.expira_en:
        solicitud.estado = 'EXPIRADO'
        solicitud.save(update_fields=['estado', 'actualizado_en'])

    tipo_labels = {
        'CONTRATO': 'Contrato Laboral', 'ANEXO_40H': 'Anexo Ley 40 Horas',
        'AMONESTACION': 'Carta de Amonestación', 'DESPIDO': 'Carta de Despido',
        'CONSTANCIA': 'Constancia Laboral', 'ANEXO_CONTRATO': 'Anexo de Contrato',
        'LIQUIDACION': 'Liquidación de Sueldo', 'VACACION': 'Comprobante de Vacaciones',
        'FINIQUITO': 'Finiquito de Término',
    }
    empleado = solicitud.empleado
    empresa  = solicitud.empresa

    return Response({
        'estado': solicitud.estado,
        'tipo_documento': solicitud.tipo_documento,
        'tipo_documento_label': tipo_labels.get(solicitud.tipo_documento, solicitud.tipo_documento),
        'empresa_nombre': empresa.nombre_legal,
        'trabajador_nombre': f"{empleado.nombres} {empleado.apellido_paterno}",
        'email_firmante_enmascarado': _enmascarar_email(solicitud.email_firmante),
        'expira_en': solicitud.expira_en.isoformat(),
        'ya_verificado': solicitud.sesion_token_trabajador is not None,
    })


@api_view(['POST'])
@permission_classes([AllowAny])
def firma_publica_solicitar_otp(request, token):
    """
    Genera un código OTP de 6 dígitos y lo envía al email del trabajador.
    Limita a 1 solicitud por minuto para evitar spam.
    """
    try:
        solicitud = SolicitudFirma.objects.get(token=token)
    except SolicitudFirma.DoesNotExist:
        return Response({'error': 'Solicitud de firma no encontrada.'}, status=404)

    if solicitud.estado != 'PENDIENTE':
        return Response({'error': 'Esta solicitud no está pendiente de firma.'}, status=400)

    if timezone.now() > solicitud.expira_en:
        solicitud.estado = 'EXPIRADO'
        solicitud.save(update_fields=['estado', 'actualizado_en'])
        return Response({'error': 'El enlace de firma ha expirado.'}, status=410)

    # Anti-spam: no permitir más de un OTP por minuto
    ultimo_otp = OTPFirma.objects.filter(solicitud=solicitud).order_by('-creado_en').first()
    if ultimo_otp:
        segundos_transcurridos = (timezone.now() - ultimo_otp.creado_en).total_seconds()
        if segundos_transcurridos < 60:
            espera = int(60 - segundos_transcurridos)
            return Response(
                {'error': f'Debes esperar {espera} segundo(s) antes de solicitar un nuevo código.'},
                status=429
            )

    # Invalidar OTPs previos no verificados
    OTPFirma.objects.filter(
        solicitud=solicitud, verificado=False
    ).update(expira_en=timezone.now())

    # Generar código de 6 dígitos
    codigo = ''.join(random.choices(string.digits, k=6))

    otp = OTPFirma.objects.create(
        solicitud=solicitud,
        codigo=codigo,
        email_destino=solicitud.email_firmante,
    )

    # Enviar email con el código
    try:
        _enviar_email_otp(otp, solicitud)
    except Exception as e:
        otp.delete()
        return Response({'error': f'No se pudo enviar el código por email: {str(e)}'}, status=500)

    return Response({
        'enviado': True,
        'email_destino': _enmascarar_email(solicitud.email_firmante),
        'expira_en_minutos': 10,
    })


@api_view(['POST'])
@permission_classes([AllowAny])
def firma_publica_verificar_otp(request, token):
    """
    Verifica el código OTP enviado al trabajador.
    Si es correcto devuelve un sesion_token que autoriza el paso de firma.
    """
    codigo_enviado = str(request.data.get('codigo', '')).strip()

    if not codigo_enviado:
        return Response({'error': 'Debes ingresar el código.'}, status=400)

    try:
        solicitud = SolicitudFirma.objects.get(token=token)
    except SolicitudFirma.DoesNotExist:
        return Response({'error': 'Solicitud de firma no encontrada.'}, status=404)

    if solicitud.estado != 'PENDIENTE':
        return Response({'error': 'Esta solicitud no está pendiente de firma.'}, status=400)

    if timezone.now() > solicitud.expira_en:
        solicitud.estado = 'EXPIRADO'
        solicitud.save(update_fields=['estado', 'actualizado_en'])
        return Response({'error': 'El enlace de firma ha expirado.'}, status=410)

    # Buscar el OTP más reciente válido
    otp = OTPFirma.objects.filter(
        solicitud=solicitud, verificado=False
    ).order_by('-creado_en').first()

    if not otp or not otp.es_valido:
        return Response(
            {'error': 'No hay un código activo. Por favor solicita uno nuevo.'},
            status=400
        )

    # Incrementar intentos antes de verificar (previene timing attacks)
    otp.intentos += 1
    otp.save(update_fields=['intentos'])

    if otp.intentos > 3:
        return Response(
            {'error': 'Código bloqueado por demasiados intentos. Solicita uno nuevo.'},
            status=400
        )

    if otp.codigo != codigo_enviado:
        restantes = 3 - otp.intentos
        msg = (f'Código incorrecto. Te quedan {restantes} intento(s).'
               if restantes > 0 else 'Código bloqueado. Solicita uno nuevo.')
        return Response({'error': msg}, status=400)

    # Código correcto — marcar OTP como verificado
    otp.verificado = True
    otp.save(update_fields=['verificado'])

    # Generar sesion_token para el paso de firma
    sesion_token = uuid_mod.uuid4()
    solicitud.sesion_token_trabajador = sesion_token
    solicitud.save(update_fields=['sesion_token_trabajador', 'actualizado_en'])

    return Response({
        'verificado': True,
        'sesion_token': str(sesion_token),
    })


def _enviar_email_otp(otp: OTPFirma, solicitud: SolicitudFirma):
    """Envía el código OTP al trabajador por email."""
    tipo_labels = {
        'CONTRATO': 'Contrato Laboral', 'ANEXO_40H': 'Anexo Ley 40 Horas',
        'AMONESTACION': 'Carta de Amonestación', 'DESPIDO': 'Carta de Despido',
        'CONSTANCIA': 'Constancia Laboral', 'ANEXO_CONTRATO': 'Anexo de Contrato',
        'LIQUIDACION': 'Liquidación de Sueldo', 'VACACION': 'Comprobante de Vacaciones',
        'FINIQUITO': 'Finiquito de Término',
    }
    tipo_label = tipo_labels.get(solicitud.tipo_documento, solicitud.tipo_documento)
    empresa_nombre = solicitud.empresa.nombre_legal
    codigo = otp.codigo

    texto_plano = (
        f"Tu código de verificación es: {codigo}\n\n"
        f"Ingresa este código en la página de firma para verificar tu identidad.\n"
        f"Válido por 10 minutos.\n\n"
        f"Si no solicitaste este código, ignora este mensaje.\n\n"
        f"Jornada40 — Sistema de Gestión Laboral"
    )
    html_body = f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;background:#f4f6f9;margin:0;padding:0;">
  <div style="max-width:520px;margin:40px auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.08);">
    <div style="background:linear-gradient(135deg,#0c1a35,#1e3a6e);padding:32px 40px;">
      <h1 style="color:#fff;margin:0;font-size:20px;font-weight:700;">Verifica tu identidad</h1>
      <p style="color:rgba(255,255,255,0.6);margin:6px 0 0;font-size:14px;">{empresa_nombre} · {tipo_label}</p>
    </div>
    <div style="padding:36px 40px;">
      <p style="color:#374151;font-size:14px;margin:0 0 24px;line-height:1.6;">
        Ingresa el siguiente código en la página de firma para verificar tu identidad:
      </p>
      <div style="text-align:center;margin:0 0 28px;">
        <div style="display:inline-block;background:#f0f4ff;border:2px dashed #2563eb;border-radius:12px;padding:20px 40px;">
          <span style="font-size:38px;font-weight:900;letter-spacing:0.3em;color:#1e3a6e;font-family:monospace;">{codigo}</span>
        </div>
        <p style="color:#6b7280;font-size:13px;margin:10px 0 0;">Válido por <strong>10 minutos</strong></p>
      </div>
      <hr style="border:none;border-top:1px solid #e5e7eb;margin:24px 0;">
      <p style="color:#9ca3af;font-size:11px;margin:0;line-height:1.6;">
        Si no solicitaste este código, puedes ignorar este mensaje con seguridad.<br>
        Firma Electrónica Simple válida bajo Ley 19.799 (Chile).
      </p>
    </div>
  </div>
</body>
</html>"""

    msg = EmailMultiAlternatives(
        subject=f"Tu código de verificación — {empresa_nombre}",
        body=texto_plano,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[otp.email_destino],
    )
    msg.attach_alternative(html_body, "text/html")
    msg.send()


# ==========================================
# FASE 8 — Procesamiento de la firma
# ==========================================

def _ip_desde_request(request) -> str:
    """Extrae la IP real del firmante considerando proxies (Railway/Vercel)."""
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR', '')
    return forwarded.split(',')[0].strip() if forwarded else request.META.get('REMOTE_ADDR', '')


def _enviar_emails_firma_completada(
    solicitud: SolicitudFirma,
    empleado,
    empresa,
    pdf_firmado_bytes: bytes,
):
    """Envía confirmación de firma al trabajador y notificación al empleador."""
    tipo_labels = {
        'CONTRATO': 'Contrato Laboral', 'ANEXO_40H': 'Anexo Ley 40 Horas',
        'AMONESTACION': 'Carta de Amonestación', 'DESPIDO': 'Carta de Despido',
        'CONSTANCIA': 'Constancia Laboral', 'ANEXO_CONTRATO': 'Anexo de Contrato',
        'LIQUIDACION': 'Liquidación de Sueldo', 'VACACION': 'Comprobante de Vacaciones',
        'FINIQUITO': 'Finiquito de Término',
    }
    tipo_label        = tipo_labels.get(solicitud.tipo_documento, solicitud.tipo_documento)
    nombre_trabajador = f"{empleado.nombres} {empleado.apellido_paterno}"
    firmado_str       = solicitud.firmado_en.strftime('%d/%m/%Y a las %H:%M') + ' UTC'
    nombre_pdf        = f"{tipo_label.replace(' ', '_')}_{empleado.rut}_firmado.pdf"

    # ── Email al trabajador ──────────────────────────────────────────────────
    texto_trabajador = (
        f"Hola {nombre_trabajador},\n\n"
        f"Tu firma electrónica simple fue registrada exitosamente el {firmado_str}.\n\n"
        f"Documento: {tipo_label}\n"
        f"Empresa: {empresa.nombre_legal}\n\n"
        f"Adjunto encontrarás una copia del documento firmado con el certificado de firma.\n\n"
        f"Jornada40 — Sistema de Gestión Laboral"
    )
    html_trabajador = f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;background:#f4f6f9;margin:0;padding:0;">
  <div style="max-width:560px;margin:40px auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.08);">
    <div style="background:linear-gradient(135deg,#0c1a35,#1e3a6e);padding:32px 40px;">
      <h1 style="color:#fff;margin:0;font-size:22px;font-weight:700;">¡Documento Firmado!</h1>
      <p style="color:rgba(255,255,255,0.6);margin:6px 0 0;font-size:14px;">{empresa.nombre_legal}</p>
    </div>
    <div style="padding:32px 40px;">
      <p style="color:#374151;font-size:15px;margin:0 0 8px;">Hola <strong>{nombre_trabajador}</strong>,</p>
      <p style="color:#374151;font-size:14px;line-height:1.6;margin:0 0 24px;">
        Tu firma electrónica simple fue registrada exitosamente.
      </p>
      <div style="background:#f0fdf4;border-left:4px solid #059669;padding:16px 20px;border-radius:6px;margin-bottom:28px;">
        <p style="margin:0;font-weight:700;color:#065f46;font-size:15px;">{tipo_label}</p>
        <p style="margin:4px 0 0;color:#6b7280;font-size:13px;">Firmado el {firmado_str}</p>
      </div>
      <p style="color:#374151;font-size:14px;line-height:1.6;margin:0 0 24px;">
        Adjunto a este correo encontrarás el documento firmado con el certificado de autenticidad.
        Guárdalo en un lugar seguro.
      </p>
      <hr style="border:none;border-top:1px solid #e5e7eb;margin:24px 0;">
      <p style="color:#9ca3af;font-size:11px;margin:0;line-height:1.6;">
        Firma Electrónica Simple válida bajo Ley N° 19.799 (Chile). Generado por Jornada40.
      </p>
    </div>
  </div>
</body>
</html>"""

    msg_trabajador = EmailMultiAlternatives(
        subject=f"Documento firmado: {tipo_label} — {empresa.nombre_legal}",
        body=texto_trabajador,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[solicitud.email_firmante],
    )
    msg_trabajador.attach_alternative(html_trabajador, "text/html")
    msg_trabajador.attach(nombre_pdf, pdf_firmado_bytes, 'application/pdf')
    msg_trabajador.send()

    # ── Email al empleador ───────────────────────────────────────────────────
    email_empleador = empresa.owner.email
    if not email_empleador:
        return

    texto_empleador = (
        f"El trabajador {nombre_trabajador} firmó el documento «{tipo_label}» "
        f"el {firmado_str}.\n\n"
        f"Empresa: {empresa.nombre_legal}\n"
        f"El documento firmado está adjunto a este correo.\n\n"
        f"Jornada40 — Sistema de Gestión Laboral"
    )
    html_empleador = f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;background:#f4f6f9;margin:0;padding:0;">
  <div style="max-width:560px;margin:40px auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.08);">
    <div style="background:linear-gradient(135deg,#0c1a35,#1e3a6e);padding:32px 40px;">
      <h1 style="color:#fff;margin:0;font-size:22px;font-weight:700;">Firma Recibida</h1>
      <p style="color:rgba(255,255,255,0.6);margin:6px 0 0;font-size:14px;">{empresa.nombre_legal}</p>
    </div>
    <div style="padding:32px 40px;">
      <p style="color:#374151;font-size:14px;line-height:1.6;margin:0 0 24px;">
        El trabajador <strong>{nombre_trabajador}</strong> firmó el siguiente documento:
      </p>
      <div style="background:#f0f4ff;border-left:4px solid #2563eb;padding:16px 20px;border-radius:6px;margin-bottom:28px;">
        <p style="margin:0;font-weight:700;color:#1e3a6e;font-size:15px;">{tipo_label}</p>
        <p style="margin:4px 0 0;color:#6b7280;font-size:13px;">Firmado el {firmado_str}</p>
      </div>
      <p style="color:#374151;font-size:14px;line-height:1.6;margin:0 0 0;">
        El documento firmado con certificado de autenticidad está adjunto a este correo.
      </p>
      <hr style="border:none;border-top:1px solid #e5e7eb;margin:24px 0;">
      <p style="color:#9ca3af;font-size:11px;margin:0;line-height:1.6;">
        Firma Electrónica Simple válida bajo Ley N° 19.799 (Chile). Generado por Jornada40.
      </p>
    </div>
  </div>
</body>
</html>"""

    msg_empleador = EmailMultiAlternatives(
        subject=f"Firma recibida: {nombre_trabajador} firmó «{tipo_label}»",
        body=texto_empleador,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[email_empleador],
    )
    msg_empleador.attach_alternative(html_empleador, "text/html")
    msg_empleador.attach(nombre_pdf, pdf_firmado_bytes, 'application/pdf')
    msg_empleador.send()


@api_view(['POST'])
@permission_classes([AllowAny])
def firma_publica_firmar(request, token):
    """
    Procesa la firma del trabajador:
    1. Valida sesion_token y datos de entrada
    2. Descarga PDF original de B2
    3. Genera PDF firmado con página de certificado (pdf_firma.py)
    4. Sube PDF firmado a B2 y elimina el temporal
    5. Marca SolicitudFirma como FIRMADO
    6. Envía emails de confirmación con el PDF adjunto
    """
    sesion_token    = str(request.data.get('sesion_token',    '')).strip()
    firma_trabajador = str(request.data.get('firma_trabajador', '')).strip()

    if not sesion_token:
        return Response({'error': 'Sesión inválida. Vuelve a verificar tu identidad.'}, status=400)
    if not firma_trabajador:
        return Response({'error': 'Debes dibujar tu firma antes de continuar.'}, status=400)
    if not firma_trabajador.startswith('data:image/'):
        return Response({'error': 'Formato de firma inválido.'}, status=400)
    if len(firma_trabajador) > 500_000:
        return Response({'error': 'La imagen de firma es demasiado grande.'}, status=400)

    try:
        solicitud = SolicitudFirma.objects.select_related(
            'empleado', 'empresa', 'empresa__owner'
        ).get(token=token)
    except SolicitudFirma.DoesNotExist:
        return Response({'error': 'Solicitud de firma no encontrada.'}, status=404)

    # ── Validaciones de estado ──────────────────────────────────────────────
    if solicitud.estado != 'PENDIENTE':
        return Response(
            {'error': f'Esta solicitud ya no está pendiente ({solicitud.get_estado_display()}).'},
            status=400,
        )

    if timezone.now() > solicitud.expira_en:
        solicitud.estado = 'EXPIRADO'
        solicitud.save(update_fields=['estado', 'actualizado_en'])
        return Response({'error': 'El enlace de firma ha expirado.'}, status=410)

    if (not solicitud.sesion_token_trabajador
            or str(solicitud.sesion_token_trabajador) != sesion_token):
        return Response(
            {'error': 'Sesión inválida. Vuelve a verificar tu identidad con el código OTP.'},
            status=403,
        )

    if not solicitud.b2_key_temporal:
        return Response(
            {'error': 'No se encontró el PDF del documento. Contacta al empleador.'},
            status=400,
        )

    # ── Descargar PDF original de B2 ────────────────────────────────────────
    try:
        pdf_original_bytes = b2_client.descargar_documento(solicitud.b2_key_temporal)
    except Exception as exc:
        return Response({'error': f'Error al obtener el documento: {exc}'}, status=500)

    # ── Generar PDF firmado con certificado ─────────────────────────────────
    empleado = solicitud.empleado
    empresa  = solicitud.empresa
    tipo_labels = {
        'CONTRATO': 'Contrato Laboral', 'ANEXO_40H': 'Anexo Ley 40 Horas',
        'AMONESTACION': 'Carta de Amonestación', 'DESPIDO': 'Carta de Despido',
        'CONSTANCIA': 'Constancia Laboral', 'ANEXO_CONTRATO': 'Anexo de Contrato',
        'LIQUIDACION': 'Liquidación de Sueldo', 'VACACION': 'Comprobante de Vacaciones',
        'FINIQUITO': 'Finiquito de Término',
    }
    firmado_en  = timezone.now()
    ip_firmante = _ip_desde_request(request)

    try:
        from .pdf_firma import agregar_certificado_firma
        pdf_firmado_bytes = agregar_certificado_firma(
            pdf_original_bytes    = pdf_original_bytes,
            tipo_documento_label  = tipo_labels.get(solicitud.tipo_documento, solicitud.tipo_documento),
            empresa_nombre        = empresa.nombre_legal,
            empresa_rut           = empresa.rut,
            firmante_nombre       = empresa.firma_firmante_nombre or empresa.representante_legal or '',
            firmante_cargo        = empresa.firma_firmante_cargo or 'Representante Legal',
            firma_empleador_b64   = empresa.firma_imagen or '',
            trabajador_nombre     = f"{empleado.nombres} {empleado.apellido_paterno}",
            trabajador_rut        = empleado.rut,
            firma_trabajador_b64  = firma_trabajador,
            token                 = str(solicitud.token),
            firmado_en            = firmado_en,
            ip_firmante           = ip_firmante,
            email_firmante        = solicitud.email_firmante,
        )
    except Exception as exc:
        return Response({'error': f'Error al generar el documento firmado: {exc}'}, status=500)

    # ── Subir PDF firmado a B2 ──────────────────────────────────────────────
    key_firmado = b2_client.key_firmado(
        empresa_id=empresa.id,
        uuid=str(solicitud.token),
        year=firmado_en.year,
        month=firmado_en.month,
    )
    try:
        b2_client.subir_documento(pdf_firmado_bytes, key_firmado)
    except Exception as exc:
        return Response({'error': f'Error al guardar el documento firmado: {exc}'}, status=500)

    # Eliminar PDF temporal (no crítico)
    b2_client.eliminar_documento(solicitud.b2_key_temporal)

    # ── Actualizar SolicitudFirma ───────────────────────────────────────────
    solicitud.estado                 = 'FIRMADO'
    solicitud.firmado_en             = firmado_en
    solicitud.ip_firmante            = ip_firmante or None
    solicitud.firma_trabajador_imagen = firma_trabajador
    solicitud.b2_key_firmado         = key_firmado
    solicitud.sesion_token_trabajador = None   # invalidar sesión
    solicitud.save(update_fields=[
        'estado', 'firmado_en', 'ip_firmante',
        'firma_trabajador_imagen', 'b2_key_firmado',
        'sesion_token_trabajador', 'actualizado_en',
    ])

    # ── Emails de confirmación (no críticos) ────────────────────────────────
    try:
        _enviar_emails_firma_completada(solicitud, empleado, empresa, pdf_firmado_bytes)
    except Exception:
        pass

    return Response({'firmado': True, 'firmado_en': firmado_en.isoformat()})


# ============================================================
# FASE 9 — Previsualización y rechazo por parte del trabajador
# ============================================================

_TIPO_LABELS_PUBLICO = {
    'CONTRATO':       'Contrato Laboral',
    'ANEXO_40H':      'Anexo Ley 40 Horas',
    'AMONESTACION':   'Carta de Amonestación',
    'DESPIDO':        'Carta de Despido',
    'CONSTANCIA':     'Constancia Laboral',
    'ANEXO_CONTRATO': 'Anexo de Contrato',
    'LIQUIDACION':    'Liquidación de Sueldo',
    'VACACION':       'Comprobante de Vacaciones',
    'FINIQUITO':      'Finiquito de Término',
}


@api_view(['GET'])
@permission_classes([AllowAny])
def firma_publica_documento(request, token):
    """
    Retorna el PDF del documento para que el trabajador lo revise antes de firmar.
    Solo requiere el token — ver el documento no constituye firma ni compromiso.
    Si la solicitud ya fue firmada, retorna el PDF firmado con certificado.
    """
    try:
        solicitud = SolicitudFirma.objects.select_related('empleado').get(token=token)
    except SolicitudFirma.DoesNotExist:
        return HttpResponse(status=404)

    if solicitud.estado in ('CANCELADO', 'EXPIRADO'):
        return HttpResponse(status=410)

    # Elegir qué versión del PDF servir
    if solicitud.estado == 'FIRMADO' and solicitud.b2_key_firmado:
        b2_key = solicitud.b2_key_firmado
    elif solicitud.b2_key_temporal:
        b2_key = solicitud.b2_key_temporal
    else:
        return HttpResponse(status=404)

    try:
        pdf_bytes = b2_client.descargar_documento(b2_key)
    except Exception as exc:
        return Response({'error': f'Error al obtener el documento: {exc}'}, status=500)

    tipo_label = _TIPO_LABELS_PUBLICO.get(solicitud.tipo_documento, solicitud.tipo_documento)
    apellido   = solicitud.empleado.apellido_paterno.replace(' ', '_')
    filename   = f"{tipo_label.replace(' ', '_')}_{apellido}.pdf"

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    response['Content-Length']      = len(pdf_bytes)
    return response


@api_view(['POST'])
@permission_classes([AllowAny])
def firma_publica_rechazar(request, token):
    """
    El trabajador rechaza el documento tras verificar su identidad con OTP.
    Requiere sesion_token válido. Cambia estado a RECHAZADO y notifica al empleador.
    """
    sesion_token = str(request.data.get('sesion_token', '')).strip()
    motivo       = str(request.data.get('motivo', '')).strip()

    if not sesion_token:
        return Response({'error': 'Sesión inválida. Vuelve a verificar tu identidad.'}, status=400)

    try:
        solicitud = SolicitudFirma.objects.select_related(
            'empleado', 'empresa', 'empresa__owner'
        ).get(token=token)
    except SolicitudFirma.DoesNotExist:
        return Response({'error': 'Solicitud no encontrada.'}, status=404)

    if solicitud.estado != 'PENDIENTE':
        return Response(
            {'error': f'Esta solicitud ya no está pendiente ({solicitud.get_estado_display()}).'},
            status=400,
        )

    if timezone.now() > solicitud.expira_en:
        solicitud.estado = 'EXPIRADO'
        solicitud.save(update_fields=['estado', 'actualizado_en'])
        return Response({'error': 'El enlace de firma ha expirado.'}, status=410)

    if (not solicitud.sesion_token_trabajador
            or str(solicitud.sesion_token_trabajador) != sesion_token):
        return Response(
            {'error': 'Sesión inválida. Vuelve a verificar tu identidad con el código OTP.'},
            status=403,
        )

    solicitud.estado = 'RECHAZADO'
    solicitud.motivo_rechazo = motivo
    solicitud.save(update_fields=['estado', 'motivo_rechazo', 'actualizado_en'])

    try:
        _notificar_rechazo_empleador(solicitud, motivo)
    except Exception:
        pass  # el rechazo ya fue registrado; el email es no-crítico

    return Response({'rechazado': True})


def _notificar_rechazo_empleador(solicitud: SolicitudFirma, motivo: str):
    """Envía un email al empleador informando que el trabajador rechazó el documento."""
    tipo_label      = _TIPO_LABELS_PUBLICO.get(solicitud.tipo_documento, solicitud.tipo_documento)
    empleado        = solicitud.empleado
    empresa         = solicitud.empresa
    nombre_trabajador = f"{empleado.nombres} {empleado.apellido_paterno}"
    email_empleador = empresa.owner.email
    if not email_empleador:
        return

    _motivo_seg = _esc(motivo) if motivo else ''
    motivo_bloque = (
        f"<div style='background:#fff3cd;border-left:4px solid #f59e0b;padding:12px 16px;"
        f"border-radius:6px;margin:20px 0;'>"
        f"<p style='margin:0;font-size:13px;color:#92400e;font-weight:600;'>Motivo indicado</p>"
        f"<p style='margin:4px 0 0;font-size:14px;color:#374151;'>{_motivo_seg}</p></div>"
        if motivo else ""
    )
    motivo_texto = f"\n\nMotivo indicado por el trabajador: {motivo}" if motivo else ""

    texto_plano = (
        f"El trabajador {nombre_trabajador} rechazó la firma del documento «{tipo_label}».\n"
        f"Empresa: {empresa.nombre_legal}{motivo_texto}\n\n"
        f"Revisa el documento y comunícate con el trabajador para resolver el inconveniente.\n\n"
        f"Jornada40 — Sistema de Gestión Laboral"
    )
    html_body = f"""<!DOCTYPE html>
<html lang="es">
<head><meta charset="UTF-8"></head>
<body style="font-family:Arial,sans-serif;background:#f4f6f9;margin:0;padding:0;">
  <div style="max-width:560px;margin:40px auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.08);">
    <div style="background:linear-gradient(135deg,#7f1d1d,#991b1b);padding:32px 40px;">
      <h1 style="color:#fff;margin:0;font-size:22px;font-weight:700;">Documento Rechazado</h1>
      <p style="color:rgba(255,255,255,0.7);margin:6px 0 0;font-size:14px;">{empresa.nombre_legal}</p>
    </div>
    <div style="padding:32px 40px;">
      <p style="color:#374151;font-size:14px;line-height:1.6;margin:0 0 16px;">
        El trabajador <strong>{nombre_trabajador}</strong> rechazó la firma del siguiente documento:
      </p>
      <div style="background:#fef2f2;border-left:4px solid #ef4444;padding:16px 20px;border-radius:6px;margin-bottom:20px;">
        <p style="margin:0;font-weight:700;color:#991b1b;font-size:15px;">{tipo_label}</p>
        <p style="margin:4px 0 0;color:#6b7280;font-size:13px;">{empresa.nombre_legal}</p>
      </div>
      {motivo_bloque}
      <p style="color:#374151;font-size:14px;line-height:1.6;margin:0;">
        Comunícate con el trabajador para revisar el documento y volver a enviarlo una vez corregido.
      </p>
      <hr style="border:none;border-top:1px solid #e5e7eb;margin:24px 0;">
      <p style="color:#9ca3af;font-size:11px;margin:0;line-height:1.6;">
        Firma Electrónica Simple · Ley N° 19.799 · Jornada40
      </p>
    </div>
  </div>
</body>
</html>"""

    msg = EmailMultiAlternatives(
        subject=f"Documento rechazado: {nombre_trabajador} rechazó «{tipo_label}»",
        body=texto_plano,
        from_email=settings.DEFAULT_FROM_EMAIL,
        to=[email_empleador],
    )
    msg.attach_alternative(html_body, "text/html")
    msg.send()